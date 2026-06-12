"""
Excel Model Auditor - pywin32 COM Automation
Performs comprehensive auditing of Excel financial models.

Usage:
    python excel_auditor.py <file_path> [options]

Options:
    --sheets "Sheet1,Sheet2"     Target specific sheets (default: all)
    --checks "formula,format,..."  Which checks to run (default: all)
    --output-format xlsx|md|html   Report format (default: xlsx)
    --output-path <path>           Output location (default: same as input)
    --severity critical|warning|info|all  Minimum severity (default: all)
"""

import sys
import os
import json
import re
import argparse
import time
from datetime import datetime
from collections import defaultdict, Counter
from pathlib import Path

try:
    import win32com.client as win32
    import pythoncom
except ImportError:
    print("ERROR: pywin32 is required. Install with: pip install pywin32")
    sys.exit(1)

# ============================================================================
# Constants
# ============================================================================

ALL_CHECKS = ["formula", "format", "validation", "hyperlinks", "names", "vba", "pq", "pp"]

SEVERITY_ORDER = {"Critical": 0, "Warning": 1, "Info": 2}

# Display values for the Severity column (audit_standards.md section 3).
# Emoji are written only to report FILES (always UTF-8), never to the console,
# which may be cp1252 on Windows.
SEVERITY_DISPLAY = {
    "Critical": "\U0001F534 Critical",   # red circle
    "Warning": "⚠️ Warning",   # warning sign
    "Info": "\U0001F7E1 Info",           # yellow circle
}

# Unified specialist findings columns (audit_standards.md section 1).
REPORT_COLUMNS = ["Sheet", "Cell Reference", "Description of Location",
                  "Severity", "Category", "Description of Issue"]

SEMANTIC_RULES = {
    "R001": {
        "name": "Total/Subtotal",
        "label_patterns": [r"\btotal\b", r"\bsubtotal\b", r"\bsum\b", r"\bgrand\s*total\b"],
        "expected_functions": ["SUM", "SUBTOTAL", "AGGREGATE"],
        "expected_operators": ["+"],
        "flag_operators": ["*", "/"],
        "severity": "Critical",
        "comment": "Label indicates a total/sum but formula does not aggregate."
    },
    "R002": {
        "name": "Per Unit/Average",
        "label_patterns": [r"\bper\s+unit\b", r"\bper\s+item\b", r"\baverage\b", r"\bavg\b",
                          r"\bmean\b", r"\bunit\s+price\b", r"\bunit\s+cost\b", r"\bper\s+capita\b"],
        "expected_functions": ["AVERAGE", "AVERAGEIF", "AVERAGEIFS"],
        "expected_operators": ["/"],
        "flag_operators": ["*"],
        "flag_condition": "multiplication_without_division",
        "severity": "Critical",
        "comment": "Label suggests per-unit/average calculation but formula does not divide."
    },
    "R003": {
        "name": "Percentage/Ratio",
        "label_patterns": [r"%", r"\bpercent", r"\bratio\b", r"\bproportion\b",
                          r"\bmargin\b", r"\byield\b"],
        "expected_operators": ["/"],
        "severity": "Warning",
        "comment": "Label indicates a percentage/ratio but formula does not compute a proportion."
    },
    "R004": {
        "name": "Variance/Difference",
        "label_patterns": [r"\bvariance\b", r"\bdifference\b", r"\bdelta\b",
                          r"\bchange\b", r"\bdeviation\b"],
        "expected_operators": ["-"],
        "severity": "Warning",
        "comment": "Label suggests a variance/difference but formula does not subtract."
    },
    "R005": {
        "name": "Growth Rate",
        "label_patterns": [r"\bgrowth\b", r"\byoy\b", r"\by-o-y\b", r"\bmom\b",
                          r"\bcagr\b", r"\bgrowth\s+rate\b"],
        "expected_operators": ["-", "/"],
        "severity": "Warning",
        "comment": "Label suggests a growth calculation but formula does not compare periods."
    },
    "R006": {
        "name": "Count",
        "label_patterns": [r"\bcount\b", r"\bnumber\s+of\b", r"\b#\s*of\b",
                          r"\bqty\b", r"\bquantity\b", r"\bheadcount\b"],
        "expected_functions": ["COUNT", "COUNTA", "COUNTIF", "COUNTIFS", "ROWS", "COLUMNS"],
        "severity": "Warning",
        "comment": "Label indicates a count but formula performs arithmetic rather than counting."
    },
    "R007": {
        "name": "Cumulative/YTD",
        "label_patterns": [r"\bcumulative\b", r"\brunning\s+total\b", r"\bytd\b",
                          r"\byear\s+to\s+date\b", r"\baccumulated\b"],
        "severity": "Warning",
        "comment": "Label suggests cumulative calculation but formula does not accumulate."
    },
    "R008": {
        "name": "Max/Min",
        "label_patterns": [r"\bmax\b", r"\bmaximum\b", r"\bmin\b", r"\bminimum\b",
                          r"\bhighest\b", r"\blowest\b", r"\bpeak\b"],
        "expected_functions": ["MAX", "MIN", "LARGE", "SMALL", "MAXIFS", "MINIFS"],
        "severity": "Info",
        "comment": "Label suggests a max/min operation but formula does not use MAX/MIN."
    },
    "R009": {
        "name": "Weighted Average",
        "label_patterns": [r"\bweighted\b", r"\bwacc\b", r"\bblended\b", r"\bcomposite\b"],
        "expected_functions": ["SUMPRODUCT"],
        "severity": "Warning",
        "comment": "Label suggests weighted calculation but formula does not apply weights."
    },
    "R010": {
        "name": "Balance Roll-forward",
        "label_patterns": [r"\bopening\s+balance\b", r"\bclosing\s+balance\b",
                          r"\bending\s+balance\b", r"\bbeginning\s+balance\b"],
        "severity": "Warning",
        "comment": "Label indicates a balance but formula does not follow roll-forward pattern."
    },
    "R011": {
        "name": "Net/After Deductions",
        "label_patterns": [r"\bnet\b", r"\bafter\s+tax\b", r"\bpost[\s-]tax\b"],
        "expected_operators": ["-"],
        "severity": "Warning",
        "comment": "Label suggests a net figure but formula does not subtract deductions."
    },
}


# ============================================================================
# Finding class
# ============================================================================

class Finding:
    """One audit finding. Report emission follows audit_standards.md:
    no per-finding ID (IDs are assigned at consolidation time), and the
    Issue + Comment pair merges into a single Description of Issue."""

    def __init__(self, sheet, location, cell_ref, issue, comment, severity="Warning", category="General"):
        self.sheet = sheet
        self.location = location
        self.cell_ref = cell_ref
        self.issue = issue
        self.comment = comment
        self.severity = severity
        self.category = category

    def description_of_issue(self):
        """Merge Issue + Comment per audit_standards.md section 1."""
        issue = str(self.issue).strip()
        comment = str(self.comment).strip() if self.comment else ""
        if comment:
            return f"{issue} — Recommend: {comment}"
        return issue

    def to_dict(self):
        """Row for the report writers — unified specialist columns."""
        return {
            "Sheet": self.sheet,
            "Cell Reference": self.cell_ref,
            "Description of Location": self.location,
            "Severity": self.severity,
            "Category": self.category,
            "Description of Issue": self.description_of_issue(),
        }

    def to_interchange(self):
        """Findings JSON record (audit_standards.md section 5)."""
        return {
            "agent": "model-auditor",
            "sheet": self.sheet,
            "cells": [self.cell_ref],
            "location": self.location,
            "severity": self.severity.lower(),
            "category": self.category,
            "description": self.description_of_issue(),
            "r1c1_expected": None,
            "r1c1_actual": None,
        }


# ============================================================================
# Excel Connection Manager
# ============================================================================

class ExcelConnection:
    """Manages COM connection to Excel, auto-detecting or opening workbooks."""

    def __init__(self, file_path):
        self.file_path = os.path.abspath(file_path)
        self.excel = None
        self.workbook = None
        self.opened_by_us = False
        self.excel_started_by_us = False

    def connect(self):
        """Auto-detect if workbook is open, otherwise open it."""
        pythoncom.CoInitialize()

        # Try to connect to running Excel instance
        try:
            self.excel = win32.GetActiveObject("Excel.Application")
            print(f"[INFO] Connected to running Excel instance.")

            # Check if the target workbook is already open
            for wb in self.excel.Workbooks:
                if os.path.abspath(wb.FullName).lower() == self.file_path.lower():
                    self.workbook = wb
                    print(f"[INFO] Workbook already open: {wb.Name}")
                    return
        except Exception:
            # No running Excel instance
            print("[INFO] No running Excel instance found. Starting Excel...")
            self.excel = win32.Dispatch("Excel.Application")
            self.excel.Visible = False
            self.excel.DisplayAlerts = False
            self.excel_started_by_us = True

        # Open the workbook
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"File not found: {self.file_path}")

        print(f"[INFO] Opening workbook: {self.file_path}")
        self.workbook = self.excel.Workbooks.Open(self.file_path, ReadOnly=True)
        self.opened_by_us = True

    def disconnect(self):
        """Clean up — close workbook if we opened it, quit Excel if we started it."""
        try:
            if self.opened_by_us and self.workbook:
                self.workbook.Close(SaveChanges=False)
                print("[INFO] Closed workbook.")
            if self.excel_started_by_us and self.excel:
                self.excel.Quit()
                print("[INFO] Quit Excel.")
        except Exception as e:
            print(f"[WARN] Error during cleanup: {e}")
        finally:
            pythoncom.CoUninitialize()

    def get_sheets(self, target_sheets=None):
        """Return list of worksheet COM objects to audit."""
        sheets = []
        for ws in self.workbook.Worksheets:
            if target_sheets is None or ws.Name in target_sheets:
                sheets.append(ws)
        return sheets


# ============================================================================
# Audit Check: Formula Consistency & Semantic Logic
# ============================================================================

class FormulaChecker:
    """Check formula consistency and semantic correctness."""

    def __init__(self):
        self.findings = []
        self.ambiguous = []  # For AI review pass

    def check_sheet(self, ws):
        """Run all formula checks on a worksheet."""
        sheet_name = ws.Name
        print(f"  [Formula] Checking sheet: {sheet_name}")

        used_range = ws.UsedRange
        if used_range is None:
            return

        row_count = used_range.Rows.Count
        col_count = used_range.Columns.Count
        start_row = used_range.Row
        start_col = used_range.Column

        # Read formulas and values in bulk for performance
        if row_count > 1 and col_count > 1:
            formulas = used_range.Formula
            values = used_range.Value
        elif row_count == 1 and col_count == 1:
            formulas = [[used_range.Formula]]
            values = [[used_range.Value]]
        else:
            try:
                formulas = used_range.Formula
                values = used_range.Value
            except Exception:
                return

        if formulas is None:
            return

        # Convert to list of lists if needed
        if not isinstance(formulas, tuple):
            formulas = [[formulas]]
        if not isinstance(values, tuple):
            values = [[values]]

        # Check column consistency (formulas in each column should follow patterns)
        self._check_column_consistency(ws, sheet_name, formulas, values,
                                       start_row, start_col, row_count, col_count)

        # Check row consistency
        self._check_row_consistency(ws, sheet_name, formulas, values,
                                    start_row, start_col, row_count, col_count)

        # Semantic label checks
        self._check_semantic_labels(ws, sheet_name, formulas, values,
                                    start_row, start_col, row_count, col_count)

        # Check for hardcoded values in formula ranges
        self._check_hardcoded_in_formula_ranges(ws, sheet_name, formulas,
                                                 start_row, start_col, row_count, col_count)

    def _get_cell_ref(self, row, col):
        """Convert row/col numbers to Excel cell reference."""
        col_letter = ""
        c = col
        while c > 0:
            c, remainder = divmod(c - 1, 26)
            col_letter = chr(65 + remainder) + col_letter
        return f"{col_letter}{row}"

    def _get_formula_pattern(self, formula):
        """Extract a normalized pattern from a formula for comparison."""
        if not formula or not str(formula).startswith("="):
            return None
        # Replace cell references with pattern markers
        pattern = re.sub(r'\$?[A-Z]+\$?\d+', 'REF', str(formula))
        # Replace numbers with NUM
        pattern = re.sub(r'\b\d+\.?\d*\b', 'NUM', pattern)
        return pattern

    def _check_column_consistency(self, ws, sheet_name, formulas, values,
                                   start_row, start_col, row_count, col_count):
        """Check that formulas in each column follow a consistent pattern."""
        for c in range(col_count):
            col_formulas = []
            for r in range(row_count):
                f = formulas[r][c] if isinstance(formulas[r], tuple) else formulas[r]
                if isinstance(f, str) and f.startswith("="):
                    pattern = self._get_formula_pattern(f)
                    col_formulas.append((r, f, pattern))

            if len(col_formulas) < 3:
                continue

            # Find the dominant pattern
            patterns = [x[2] for x in col_formulas]
            pattern_counts = Counter(patterns)
            dominant_pattern, dominant_count = pattern_counts.most_common(1)[0]

            if dominant_count < len(col_formulas) * 0.7:
                continue  # No clear dominant pattern

            # Flag outliers
            for r, formula, pattern in col_formulas:
                if pattern != dominant_pattern:
                    actual_row = start_row + r
                    actual_col = start_col + c
                    cell_ref = self._get_cell_ref(actual_row, actual_col)

                    # Get location context
                    row_label = self._get_row_label(ws, actual_row, start_col, c)

                    self.findings.append(Finding(
                        sheet=sheet_name,
                        location=row_label or f"Row {actual_row}",
                        cell_ref=cell_ref,
                        issue=f"Formula breaks column pattern. Expected pattern like majority of column.",
                        comment=f"Formula: {formula} | Dominant pattern has {dominant_count}/{len(col_formulas)} cells.",
                        severity="Critical",
                        category="Formula Consistency"
                    ))

    def _check_row_consistency(self, ws, sheet_name, formulas, values,
                                start_row, start_col, row_count, col_count):
        """Check that formulas in each row follow a consistent pattern."""
        for r in range(row_count):
            row_formulas = []
            for c in range(col_count):
                f = formulas[r][c] if isinstance(formulas[r], tuple) else formulas[r]
                if isinstance(f, str) and f.startswith("="):
                    pattern = self._get_formula_pattern(f)
                    row_formulas.append((c, f, pattern))

            if len(row_formulas) < 3:
                continue

            patterns = [x[2] for x in row_formulas]
            pattern_counts = Counter(patterns)
            dominant_pattern, dominant_count = pattern_counts.most_common(1)[0]

            if dominant_count < len(row_formulas) * 0.7:
                continue

            for c, formula, pattern in row_formulas:
                if pattern != dominant_pattern:
                    actual_row = start_row + r
                    actual_col = start_col + c
                    cell_ref = self._get_cell_ref(actual_row, actual_col)

                    self.findings.append(Finding(
                        sheet=sheet_name,
                        location=f"Row {actual_row}",
                        cell_ref=cell_ref,
                        issue=f"Formula breaks row pattern. Expected pattern like majority of row.",
                        comment=f"Formula: {formula} | Dominant pattern has {dominant_count}/{len(row_formulas)} cells.",
                        severity="Critical",
                        category="Formula Consistency"
                    ))

    def _get_row_label(self, ws, row, start_col, current_col_offset):
        """Try to find a label for the given row (check columns to the left)."""
        for c in range(start_col, start_col + min(current_col_offset, 3)):
            try:
                val = ws.Cells(row, c).Value
                if val and isinstance(val, str) and len(val.strip()) > 0:
                    return val.strip()
            except Exception:
                pass
        return None

    def _get_col_header(self, ws, col, start_row, current_row_offset):
        """Try to find a column header (check rows above)."""
        for r in range(max(1, start_row), start_row + min(current_row_offset, 3)):
            try:
                val = ws.Cells(r, col).Value
                if val and isinstance(val, str) and len(val.strip()) > 0:
                    return val.strip()
            except Exception:
                pass
        return None

    def _check_semantic_labels(self, ws, sheet_name, formulas, values,
                                start_row, start_col, row_count, col_count):
        """Check if formulas match their row/column label semantics."""
        for r in range(row_count):
            for c in range(col_count):
                f = formulas[r][c] if isinstance(formulas[r], tuple) else formulas[r]
                if not isinstance(f, str) or not f.startswith("="):
                    continue

                actual_row = start_row + r
                actual_col = start_col + c

                row_label = self._get_row_label(ws, actual_row, start_col, c)
                col_header = self._get_col_header(ws, actual_col, start_row, r)

                combined_label = " ".join(filter(None, [row_label, col_header])).lower()
                if not combined_label:
                    continue

                formula_str = str(f).upper()
                matched_rule = False

                for rule_id, rule in SEMANTIC_RULES.items():
                    # Check if label matches any pattern
                    label_match = any(re.search(p, combined_label, re.IGNORECASE)
                                     for p in rule["label_patterns"])
                    if not label_match:
                        continue

                    # Check if formula meets expectations
                    has_expected = False

                    # Check expected functions
                    if "expected_functions" in rule:
                        for func in rule["expected_functions"]:
                            if func in formula_str:
                                has_expected = True
                                break

                    # Check expected operators
                    if not has_expected and "expected_operators" in rule:
                        for op in rule["expected_operators"]:
                            if op in formula_str:
                                has_expected = True
                                break

                    # Special condition: multiplication without division
                    if rule.get("flag_condition") == "multiplication_without_division":
                        if "*" in formula_str and "/" not in formula_str:
                            has_expected = False
                        elif "/" in formula_str:
                            has_expected = True

                    if not has_expected:
                        cell_ref = self._get_cell_ref(actual_row, actual_col)
                        self.findings.append(Finding(
                            sheet=sheet_name,
                            location=row_label or f"Row {actual_row}",
                            cell_ref=cell_ref,
                            issue=f"Semantic mismatch (Rule {rule_id}: {rule['name']}): {rule['comment']}",
                            comment=f"Formula: {f} | Row label: '{row_label}' | Column: '{col_header}'",
                            severity=rule["severity"],
                            category="Semantic Logic"
                        ))
                        matched_rule = True
                        break

                # If no rule matched but formula has complex characteristics, flag for AI review
                if not matched_rule and combined_label and len(formula_str) > 5:
                    ops = set(re.findall(r'[+\-*/]', formula_str.replace("E+", "").replace("E-", "")))
                    if len(ops) >= 2 or "INDIRECT" in formula_str or "OFFSET" in formula_str:
                        self.ambiguous.append({
                            "sheet": sheet_name,
                            "cell_ref": self._get_cell_ref(actual_row, actual_col),
                            "formula": str(f),
                            "row_label": row_label,
                            "col_header": col_header,
                            "cell_value": values[r][c] if isinstance(values[r], tuple) else values[r]
                        })

    def _check_hardcoded_in_formula_ranges(self, ws, sheet_name, formulas,
                                            start_row, start_col, row_count, col_count):
        """Detect hardcoded values within ranges that otherwise contain formulas."""
        for c in range(col_count):
            formula_count = 0
            hardcoded = []
            for r in range(row_count):
                f = formulas[r][c] if isinstance(formulas[r], tuple) else formulas[r]
                if isinstance(f, str) and f.startswith("="):
                    formula_count += 1
                elif f is not None and f != "" and f != 0:
                    hardcoded.append((r, f))

            # If most cells have formulas, flag the hardcoded ones
            total = formula_count + len(hardcoded)
            if total >= 3 and formula_count >= total * 0.6:
                for r, val in hardcoded:
                    actual_row = start_row + r
                    actual_col = start_col + c
                    cell_ref = self._get_cell_ref(actual_row, actual_col)
                    row_label = self._get_row_label(ws, actual_row, start_col, c)

                    self.findings.append(Finding(
                        sheet=sheet_name,
                        location=row_label or f"Row {actual_row}",
                        cell_ref=cell_ref,
                        issue=f"Hardcoded value ({val}) in a column that otherwise contains formulas.",
                        comment=f"Column has {formula_count} formulas and {len(hardcoded)} hardcoded values. Consider using a formula.",
                        severity="Warning",
                        category="Formula Consistency"
                    ))


# ============================================================================
# Audit Check: Formatting Consistency
# ============================================================================

class FormatChecker:
    """Check formatting and style consistency."""

    def __init__(self):
        self.findings = []

    def check_sheet(self, ws):
        sheet_name = ws.Name
        print(f"  [Format] Checking sheet: {sheet_name}")

        used_range = ws.UsedRange
        if used_range is None:
            return

        row_count = used_range.Rows.Count
        col_count = used_range.Columns.Count
        start_row = used_range.Row
        start_col = used_range.Column

        self._check_number_format_consistency(ws, sheet_name, start_row, start_col, row_count, col_count)
        self._check_font_consistency(ws, sheet_name, start_row, start_col, row_count, col_count)
        self._check_decimal_consistency(ws, sheet_name, start_row, start_col, row_count, col_count)

    def _check_number_format_consistency(self, ws, sheet_name, start_row, start_col, row_count, col_count):
        """Check for inconsistent number formats within columns."""
        for c in range(col_count):
            formats = {}
            for r in range(row_count):
                try:
                    cell = ws.Cells(start_row + r, start_col + c)
                    if cell.Value is not None and isinstance(cell.Value, (int, float)):
                        nf = cell.NumberFormat
                        if nf not in formats:
                            formats[nf] = []
                        formats[nf].append((start_row + r, cell.Value))
                except Exception:
                    pass

            if len(formats) > 1:
                # Find dominant format
                dominant_format = max(formats.keys(), key=lambda k: len(formats[k]))
                for nf, cells in formats.items():
                    if nf != dominant_format and len(cells) <= len(formats[dominant_format]) * 0.3:
                        for row_num, val in cells:
                            col_num = start_col + c
                            cell_ref = FormulaChecker._get_cell_ref(None, row_num, col_num)
                            self.findings.append(Finding(
                                sheet=sheet_name,
                                location=f"Row {row_num}",
                                cell_ref=cell_ref,
                                issue=f"Inconsistent number format: '{nf}' vs dominant '{dominant_format}'",
                                comment=f"Column has {len(formats)} different number formats. Standardize for consistency.",
                                severity="Warning",
                                category="Formatting"
                            ))

    def _check_font_consistency(self, ws, sheet_name, start_row, start_col, row_count, col_count):
        """Check for mixed fonts within the sheet."""
        fonts = defaultdict(int)
        sample_size = min(row_count * col_count, 500)  # Limit for performance

        checked = 0
        for r in range(row_count):
            for c in range(col_count):
                if checked >= sample_size:
                    break
                try:
                    cell = ws.Cells(start_row + r, start_col + c)
                    if cell.Value is not None:
                        font_key = f"{cell.Font.Name}|{cell.Font.Size}"
                        fonts[font_key] += 1
                        checked += 1
                except Exception:
                    pass

        if len(fonts) > 3:
            self.findings.append(Finding(
                sheet=sheet_name,
                location="Entire sheet",
                cell_ref="N/A",
                issue=f"Sheet uses {len(fonts)} different font/size combinations.",
                comment="Consider standardizing fonts for a professional appearance. "
                        f"Fonts found: {', '.join(list(fonts.keys())[:5])}",
                severity="Info",
                category="Formatting"
            ))

    def _check_decimal_consistency(self, ws, sheet_name, start_row, start_col, row_count, col_count):
        """Check for inconsistent decimal places in numeric columns."""
        for c in range(col_count):
            decimals = defaultdict(list)
            for r in range(row_count):
                try:
                    cell = ws.Cells(start_row + r, start_col + c)
                    if cell.Value is not None and isinstance(cell.Value, float):
                        val_str = str(cell.Value)
                        if "." in val_str:
                            dec_places = len(val_str.split(".")[1].rstrip("0"))
                        else:
                            dec_places = 0
                        decimals[dec_places].append(start_row + r)
                except Exception:
                    pass

            if len(decimals) > 1 and sum(len(v) for v in decimals.values()) >= 5:
                dominant = max(decimals.keys(), key=lambda k: len(decimals[k]))
                for dec, rows in decimals.items():
                    if dec != dominant and len(rows) <= 2:
                        for row_num in rows:
                            col_num = start_col + c
                            cell_ref = FormulaChecker._get_cell_ref(None, row_num, col_num)
                            self.findings.append(Finding(
                                sheet=sheet_name,
                                location=f"Row {row_num}",
                                cell_ref=cell_ref,
                                issue=f"Inconsistent decimal places: {dec} dp vs dominant {dominant} dp",
                                comment="Standardize decimal places within this column.",
                                severity="Info",
                                category="Formatting"
                            ))


# ============================================================================
# Audit Check: Hyperlinks
# ============================================================================

class HyperlinkChecker:
    """Check for broken hyperlinks."""

    def __init__(self):
        self.findings = []

    def check_sheet(self, ws):
        sheet_name = ws.Name
        print(f"  [Hyperlinks] Checking sheet: {sheet_name}")

        try:
            for hl in ws.Hyperlinks:
                try:
                    address = hl.Address or ""
                    sub_address = hl.SubAddress or ""
                    cell_ref = hl.Range.Address.replace("$", "")

                    # Check internal references
                    if sub_address and not address:
                        # Internal link — check if target sheet/range exists
                        if "!" in sub_address:
                            target_sheet = sub_address.split("!")[0].strip("'")
                            try:
                                ws.Parent.Worksheets(target_sheet)
                            except Exception:
                                self.findings.append(Finding(
                                    sheet=sheet_name,
                                    location=f"Hyperlink at {cell_ref}",
                                    cell_ref=cell_ref,
                                    issue=f"Broken internal hyperlink: target sheet '{target_sheet}' not found.",
                                    comment=f"SubAddress: {sub_address}",
                                    severity="Warning",
                                    category="Hyperlinks"
                                ))

                    # Check external file references
                    elif address and not address.startswith("http") and not address.startswith("mailto"):
                        if not os.path.exists(address):
                            self.findings.append(Finding(
                                sheet=sheet_name,
                                location=f"Hyperlink at {cell_ref}",
                                cell_ref=cell_ref,
                                issue=f"Broken file hyperlink: '{address}' not found.",
                                comment="External file reference is inaccessible.",
                                severity="Warning",
                                category="Hyperlinks"
                            ))

                    # Flag external URL links (can't verify without HTTP requests)
                    elif address.startswith("http"):
                        # Just document — URL validation would require network access
                        pass

                except Exception as e:
                    self.findings.append(Finding(
                        sheet=sheet_name,
                        location="Hyperlink",
                        cell_ref="N/A",
                        issue=f"Error reading hyperlink: {str(e)[:100]}",
                        comment="Could not fully inspect this hyperlink.",
                        severity="Info",
                        category="Hyperlinks"
                    ))
        except Exception:
            pass


# ============================================================================
# Audit Check: Name Manager
# ============================================================================

class NameChecker:
    """Check Named Ranges for issues."""

    def __init__(self):
        self.findings = []

    def check_workbook(self, workbook):
        print(f"  [Names] Checking Name Manager...")

        for name in workbook.Names:
            try:
                name_str = name.Name
                refers_to = name.RefersTo

                # Check for #REF! errors
                if "#REF!" in str(refers_to):
                    self.findings.append(Finding(
                        sheet="Workbook",
                        location=f"Named Range: {name_str}",
                        cell_ref=str(refers_to),
                        issue=f"Named range '{name_str}' contains #REF! error.",
                        comment="This name references a deleted range or sheet. Update or remove it.",
                        severity="Critical",
                        category="Name Manager"
                    ))

                # Check scope
                if "!" not in name_str:
                    # Workbook-scoped — check if there's a conflicting sheet-scoped name
                    pass

                # Check if name is used anywhere (basic check)
                # Note: Full usage detection requires searching all formulas

            except Exception as e:
                self.findings.append(Finding(
                    sheet="Workbook",
                    location=f"Named Range",
                    cell_ref="N/A",
                    issue=f"Error inspecting named range: {str(e)[:100]}",
                    comment="Could not fully analyze this named range.",
                    severity="Info",
                    category="Name Manager"
                ))


# ============================================================================
# Audit Check: Data Validation
# ============================================================================

class ValidationChecker:
    """Check Data Validation rules."""

    def __init__(self):
        self.findings = []

    def check_sheet(self, ws):
        sheet_name = ws.Name
        print(f"  [Validation] Checking sheet: {sheet_name}")

        used_range = ws.UsedRange
        if used_range is None:
            return

        # Check cells for validation
        try:
            for cell in used_range:
                try:
                    val = cell.Validation
                    if val and val.Type is not None:
                        # Check if validation formula references are valid
                        try:
                            formula1 = val.Formula1
                            if formula1 and "#REF!" in str(formula1):
                                cell_ref = cell.Address.replace("$", "")
                                self.findings.append(Finding(
                                    sheet=sheet_name,
                                    location=f"Data Validation at {cell_ref}",
                                    cell_ref=cell_ref,
                                    issue=f"Data validation rule references broken range.",
                                    comment=f"Validation formula: {formula1}",
                                    severity="Warning",
                                    category="Data Validation"
                                ))
                        except Exception:
                            pass
                except Exception:
                    pass  # Cell has no validation — this is normal
        except Exception:
            # UsedRange iteration can fail on large sheets — fall back to sampling
            pass


# ============================================================================
# Audit Check: VBA Code
# ============================================================================

class VBAChecker:
    """Review VBA code for common issues."""

    def __init__(self):
        self.findings = []

    def check_workbook(self, workbook):
        print(f"  [VBA] Checking VBA code...")

        try:
            vb_project = workbook.VBProject
        except Exception:
            self.findings.append(Finding(
                sheet="Workbook",
                location="VBA Project",
                cell_ref="N/A",
                issue="Cannot access VBA project — it may be password-protected.",
                comment="Enable 'Trust access to the VBA project object model' in Excel Trust Center, "
                        "or the project may be locked.",
                severity="Info",
                category="VBA Code"
            ))
            return

        for component in vb_project.VBComponents:
            try:
                code_module = component.CodeModule
                if code_module.CountOfLines == 0:
                    continue

                code = code_module.Lines(1, code_module.CountOfLines)
                module_name = component.Name
                self._analyze_vba(module_name, code)

            except Exception as e:
                self.findings.append(Finding(
                    sheet="Workbook",
                    location=f"VBA Module: {component.Name}",
                    cell_ref="N/A",
                    issue=f"Error reading VBA module: {str(e)[:100]}",
                    comment="Could not analyze this module.",
                    severity="Info",
                    category="VBA Code"
                ))

    def _analyze_vba(self, module_name, code):
        """Analyze VBA code for common issues."""
        lines = code.split("\n")

        # Check for Option Explicit
        has_option_explicit = any("option explicit" in line.lower() for line in lines)
        if not has_option_explicit and len(lines) > 5:
            self.findings.append(Finding(
                sheet="Workbook",
                location=f"VBA: {module_name}",
                cell_ref="N/A",
                issue="Missing 'Option Explicit' declaration.",
                comment="Add 'Option Explicit' to enforce variable declarations and catch typos.",
                severity="Warning",
                category="VBA Code"
            ))

        for i, line in enumerate(lines, 1):
            stripped = line.strip().lower()

            # Select/Activate usage
            if re.search(r'\.(select|activate)\b', stripped) and not stripped.startswith("'"):
                self.findings.append(Finding(
                    sheet="Workbook",
                    location=f"VBA: {module_name}, Line {i}",
                    cell_ref="N/A",
                    issue=f"Uses .Select/.Activate — performance anti-pattern.",
                    comment="Refactor to work with objects directly without Select/Activate.",
                    severity="Info",
                    category="VBA Code"
                ))

            # Shell commands
            if "shell" in stripped and not stripped.startswith("'"):
                self.findings.append(Finding(
                    sheet="Workbook",
                    location=f"VBA: {module_name}, Line {i}",
                    cell_ref="N/A",
                    issue="Contains Shell command — potential security concern.",
                    comment="Review this line to ensure it's not executing untrusted commands.",
                    severity="Warning",
                    category="VBA Code"
                ))

            # Hardcoded paths
            if re.search(r'["\'][a-z]:\\', stripped) and not stripped.startswith("'"):
                self.findings.append(Finding(
                    sheet="Workbook",
                    location=f"VBA: {module_name}, Line {i}",
                    cell_ref="N/A",
                    issue="Hardcoded file path detected.",
                    comment="Use environment variables or configuration for file paths.",
                    severity="Info",
                    category="VBA Code"
                ))


# ============================================================================
# Audit Check: Power Query
# ============================================================================

class PowerQueryChecker:
    """Review Power Query (M) code."""

    def __init__(self):
        self.findings = []

    def check_workbook(self, workbook):
        print(f"  [Power Query] Checking queries...")

        try:
            queries = workbook.Queries
            if queries.Count == 0:
                print("    No Power Queries found.")
                return

            for i in range(1, queries.Count + 1):
                query = queries.Item(i)
                name = query.Name
                formula = query.Formula
                self._analyze_pq(name, formula)

        except Exception as e:
            print(f"    [WARN] Could not access Power Queries: {e}")

    def _analyze_pq(self, name, formula):
        """Analyze Power Query M code for issues."""
        if not formula:
            return

        # Hardcoded file paths
        path_matches = re.findall(r'["\'][A-Za-z]:\\[^"\']+["\']', formula)
        for path in path_matches:
            self.findings.append(Finding(
                sheet="Workbook",
                location=f"Power Query: {name}",
                cell_ref="N/A",
                issue=f"Hardcoded file path in query: {path[:60]}",
                comment="Use parameters or environment-based paths for portability.",
                severity="Warning",
                category="Power Query"
            ))

        # Missing error handling
        if "try" not in formula.lower() and len(formula) > 200:
            self.findings.append(Finding(
                sheet="Workbook",
                location=f"Power Query: {name}",
                cell_ref="N/A",
                issue="No error handling (try...otherwise) detected in query.",
                comment="Add error handling for robustness, especially for external data sources.",
                severity="Info",
                category="Power Query"
            ))


# ============================================================================
# Audit Check: Power Pivot Measures
# ============================================================================

class PowerPivotChecker:
    """Review Power Pivot DAX measures."""

    def __init__(self):
        self.findings = []

    def check_workbook(self, workbook):
        print(f"  [Power Pivot] Checking DAX measures...")

        try:
            model = workbook.Model
            if model is None:
                print("    No data model found.")
                return

            # Access model tables and measures via ADO
            # Note: Direct COM access to measures can be limited
            # This is a best-effort check
            for table in model.ModelTables:
                try:
                    for measure in table.ModelMeasures:
                        name = measure.Name
                        formula = measure.Formula
                        self._analyze_dax(name, formula, table.Name)
                except Exception:
                    pass

        except Exception as e:
            print(f"    [WARN] Could not access Power Pivot model: {e}")

    def _analyze_dax(self, name, formula, table_name):
        """Analyze DAX measure for common issues."""
        if not formula:
            return

        upper = formula.upper()

        # CALCULATE without filter
        if "CALCULATE" in upper and "FILTER" not in upper and "ALL" not in upper:
            # Check if there are any filter arguments
            calc_match = re.search(r'CALCULATE\s*\([^)]+\)', upper)
            if calc_match and "," not in calc_match.group():
                self.findings.append(Finding(
                    sheet="Workbook",
                    location=f"DAX: [{table_name}]{name}",
                    cell_ref="N/A",
                    issue="CALCULATE used without explicit filter context.",
                    comment="CALCULATE without filters may not modify context as intended.",
                    severity="Info",
                    category="Power Pivot"
                ))


# ============================================================================
# Report Generator
# ============================================================================

class ReportGenerator:
    """Generate audit report in multiple formats."""

    @staticmethod
    def generate(findings, output_format, output_path, workbook_name):
        """Generate report in specified format."""
        if output_format == "xlsx":
            return ReportGenerator._generate_xlsx(findings, output_path, workbook_name)
        elif output_format == "md":
            return ReportGenerator._generate_md(findings, output_path, workbook_name)
        elif output_format == "html":
            return ReportGenerator._generate_html(findings, output_path, workbook_name)
        else:
            raise ValueError(f"Unknown format: {output_format}")

    @staticmethod
    def _generate_xlsx(findings, output_path, workbook_name):
        """Generate Excel report with separate sheets per category."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            print("openpyxl not available, falling back to Markdown.")
            return ReportGenerator._generate_md(findings, output_path, workbook_name)

        wb = Workbook()

        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"

        # Header styling
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        critical_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        info_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        severity_fills = {"Critical": critical_fill, "Warning": warning_fill, "Info": info_fill}

        # Write summary
        ws_summary["A1"] = f"Excel Model Audit Report"
        ws_summary["A1"].font = Font(bold=True, size=14)
        ws_summary["A2"] = f"Workbook: {workbook_name}"
        ws_summary["A3"] = f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws_summary["A4"] = f"Total Findings: {len(findings)}"

        # Category summary
        categories = defaultdict(lambda: {"Critical": 0, "Warning": 0, "Info": 0})
        for f in findings:
            categories[f["Category"]][f["Severity"]] += 1

        row = 6
        ws_summary.cell(row=row, column=1, value="Category").font = header_font
        ws_summary.cell(row=row, column=1).fill = header_fill
        for i, sev in enumerate(["Critical", "Warning", "Info"], 2):
            ws_summary.cell(row=row, column=i, value=sev).font = header_font
            ws_summary.cell(row=row, column=i).fill = header_fill

        for cat, counts in sorted(categories.items()):
            row += 1
            ws_summary.cell(row=row, column=1, value=cat)
            for i, sev in enumerate(["Critical", "Warning", "Info"], 2):
                cell = ws_summary.cell(row=row, column=i, value=counts[sev])
                if counts[sev] > 0:
                    cell.fill = severity_fills[sev]

        # All Findings sheet — unified specialist columns (audit_standards.md §1)
        ws_all = wb.create_sheet("All Findings")
        headers = REPORT_COLUMNS

        for c, header in enumerate(headers, 1):
            cell = ws_all.cell(row=1, column=c, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        # Sort findings by severity
        sorted_findings = sorted(findings, key=lambda x: SEVERITY_ORDER.get(x["Severity"], 99))

        for r, finding in enumerate(sorted_findings, 2):
            for c, header in enumerate(headers, 1):
                value = finding.get(header, "")
                if header == "Severity":
                    value = SEVERITY_DISPLAY.get(value, value)
                cell = ws_all.cell(row=r, column=c, value=value)
                cell.border = thin_border
                if header == "Severity":
                    cell.fill = severity_fills.get(finding["Severity"], PatternFill())

        # Auto-fit columns (approximate)
        for col in ws_all.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws_all.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

        # Save
        report_path = os.path.join(output_path, f"Audit_Report_{workbook_name}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
        wb.save(report_path)
        print(f"\n[REPORT] Excel report saved: {report_path}")
        return report_path

    @staticmethod
    def _generate_md(findings, output_path, workbook_name):
        """Generate Markdown report."""
        lines = [
            f"# Excel Model Audit Report",
            f"",
            f"**Workbook:** {workbook_name}",
            f"**Date:** {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            f"**Total Findings:** {len(findings)}",
            f"",
            f"---",
            f""
        ]

        # Summary by severity
        sev_counts = Counter(f["Severity"] for f in findings)
        lines.append("## Summary")
        lines.append("")
        lines.append(f"| Severity | Count |")
        lines.append(f"|----------|-------|")
        for sev in ["Critical", "Warning", "Info"]:
            lines.append(f"| {sev} | {sev_counts.get(sev, 0)} |")
        lines.append("")

        # Findings table — unified specialist columns (audit_standards.md §1)
        lines.append("## Findings")
        lines.append("")
        if not findings:
            lines.append("✅ No issues detected.")
        else:
            lines.append("| " + " | ".join(REPORT_COLUMNS) + " |")
            lines.append("|" + "---|" * len(REPORT_COLUMNS))

            sorted_findings = sorted(findings, key=lambda x: SEVERITY_ORDER.get(x["Severity"], 99))
            for f in sorted_findings:
                def esc(text):
                    return str(text).replace("|", "\\|").replace("\n", " ").strip()
                lines.append(f"| {esc(f['Sheet'])} | {esc(f['Cell Reference'])} | "
                            f"{esc(f['Description of Location'])} | "
                            f"{SEVERITY_DISPLAY.get(f['Severity'], f['Severity'])} | "
                            f"{esc(f['Category'])} | {esc(f['Description of Issue'])} |")

        report_path = os.path.join(output_path, f"Audit_Report_{workbook_name}_{datetime.now().strftime('%Y%m%d_%H%M')}.md")
        with open(report_path, "w", encoding="utf-8") as fp:
            fp.write("\n".join(lines))

        print(f"\n[REPORT] Markdown report saved: {report_path}")
        return report_path

    @staticmethod
    def _generate_html(findings, output_path, workbook_name):
        """Generate HTML report."""
        sorted_findings = sorted(findings, key=lambda x: SEVERITY_ORDER.get(x["Severity"], 99))

        sev_colors = {"Critical": "#FFC7CE", "Warning": "#FFEB9C", "Info": "#C6EFCE"}
        sev_counts = Counter(f["Severity"] for f in findings)

        rows_html = ""
        for f in sorted_findings:
            bg = sev_colors.get(f["Severity"], "#FFFFFF")
            rows_html += f"""<tr style="background-color:{bg}">
                <td>{f['Sheet']}</td><td>{f['Cell Reference']}</td>
                <td>{f['Description of Location']}</td>
                <td>{SEVERITY_DISPLAY.get(f['Severity'], f['Severity'])}</td>
                <td>{f['Category']}</td>
                <td>{f['Description of Issue']}</td></tr>\n"""

        html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<title>Audit Report - {workbook_name}</title>
<style>
    body {{ font-family: Calibri, Arial, sans-serif; margin: 20px; }}
    h1 {{ color: #2F5496; }}
    table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
    th {{ background: #2F5496; color: white; padding: 8px; text-align: left; }}
    td {{ border: 1px solid #ddd; padding: 6px; font-size: 13px; }}
    .summary {{ display: flex; gap: 20px; margin: 15px 0; }}
    .sev-box {{ padding: 15px 25px; border-radius: 5px; text-align: center; }}
    .critical {{ background: #FFC7CE; }}
    .warning {{ background: #FFEB9C; }}
    .info {{ background: #C6EFCE; }}
</style></head><body>
<h1>Excel Model Audit Report</h1>
<p><strong>Workbook:</strong> {workbook_name} | <strong>Date:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M')} | <strong>Total:</strong> {len(findings)} findings</p>
<div class="summary">
    <div class="sev-box critical"><strong>{sev_counts.get('Critical', 0)}</strong><br>Critical</div>
    <div class="sev-box warning"><strong>{sev_counts.get('Warning', 0)}</strong><br>Warning</div>
    <div class="sev-box info"><strong>{sev_counts.get('Info', 0)}</strong><br>Info</div>
</div>
<table>
<tr><th>Sheet</th><th>Cell Reference</th><th>Description of Location</th><th>Severity</th><th>Category</th><th>Description of Issue</th></tr>
{rows_html}
</table></body></html>"""

        report_path = os.path.join(output_path, f"Audit_Report_{workbook_name}_{datetime.now().strftime('%Y%m%d_%H%M')}.html")
        with open(report_path, "w", encoding="utf-8") as fp:
            fp.write(html)

        print(f"\n[REPORT] HTML report saved: {report_path}")
        return report_path


# ============================================================================
# Main Orchestrator
# ============================================================================

def run_audit(file_path, sheets=None, checks=None, output_format="xlsx",
              output_path=None, min_severity="all"):
    """Run the full audit pipeline."""

    if checks is None:
        checks = ALL_CHECKS
    if output_path is None:
        output_path = os.path.dirname(os.path.abspath(file_path))

    print("=" * 60)
    print("  EXCEL MODEL AUDITOR")
    print("=" * 60)
    print(f"  File: {file_path}")
    print(f"  Sheets: {sheets or 'All'}")
    print(f"  Checks: {', '.join(checks)}")
    print(f"  Output: {output_format}")
    print("=" * 60)

    # Connect to Excel
    conn = ExcelConnection(file_path)
    conn.connect()

    workbook_name = os.path.splitext(os.path.basename(file_path))[0]
    target_sheets = set(sheets) if sheets else None
    ws_list = conn.get_sheets(target_sheets)

    all_findings = []
    ambiguous_formulas = []

    try:
        # Sheet-level checks
        for ws in ws_list:
            if "formula" in checks:
                fc = FormulaChecker()
                fc.check_sheet(ws)
                all_findings.extend(fc.findings)
                ambiguous_formulas.extend(fc.ambiguous)

            if "format" in checks:
                fmt = FormatChecker()
                fmt.check_sheet(ws)
                all_findings.extend(fmt.findings)

            if "hyperlinks" in checks:
                hl = HyperlinkChecker()
                hl.check_sheet(ws)
                all_findings.extend(hl.findings)

            if "validation" in checks:
                vc = ValidationChecker()
                vc.check_sheet(ws)
                all_findings.extend(vc.findings)

        # Workbook-level checks
        wb = conn.workbook

        if "names" in checks:
            nc = NameChecker()
            nc.check_workbook(wb)
            all_findings.extend(nc.findings)

        if "vba" in checks:
            vba = VBAChecker()
            vba.check_workbook(wb)
            all_findings.extend(vba.findings)

        if "pq" in checks:
            pq = PowerQueryChecker()
            pq.check_workbook(wb)
            all_findings.extend(pq.findings)

        if "pp" in checks:
            pp = PowerPivotChecker()
            pp.check_workbook(wb)
            all_findings.extend(pp.findings)

    finally:
        conn.disconnect()

    # Filter by severity
    if min_severity != "all":
        min_order = SEVERITY_ORDER.get(min_severity.capitalize(), 99)
        all_findings = [f for f in all_findings if SEVERITY_ORDER.get(f.severity, 99) <= min_order]

    # Convert findings to report-row dicts (unified specialist columns)
    findings_data = [f.to_dict() for f in all_findings]

    # Save interchange JSON (audit_standards.md section 5 schema) so the
    # audit-manager can consolidate these findings mechanically.
    json_path = os.path.join(output_path, f"audit_findings_{workbook_name}.json")
    with open(json_path, "w", encoding="utf-8") as fp:
        json.dump({"findings": [f.to_interchange() for f in all_findings]},
                  fp, indent=2, ensure_ascii=False, default=str)
    print(f"\n[DATA] Findings JSON (audit_standards section 5): {json_path}")

    # Save ambiguous formulas for AI review
    if ambiguous_formulas:
        ambig_path = os.path.join(output_path, f"ambiguous_formulas_{workbook_name}.json")
        with open(ambig_path, "w", encoding="utf-8") as fp:
            json.dump(ambiguous_formulas, fp, indent=2, default=str)
        print(f"[DATA] Ambiguous formulas for AI review: {ambig_path}")

    # Generate report
    report_path = ReportGenerator.generate(findings_data, output_format, output_path, workbook_name)

    # Print summary
    print(f"\n{'=' * 60}")
    print(f"  AUDIT COMPLETE")
    print(f"  Total findings: {len(findings_data)}")
    sev_counts = Counter(f["Severity"] for f in findings_data)
    for sev in ["Critical", "Warning", "Info"]:
        print(f"    {sev}: {sev_counts.get(sev, 0)}")
    print(f"  Ambiguous formulas for AI review: {len(ambiguous_formulas)}")
    print(f"  Report: {report_path}")
    print(f"{'=' * 60}")

    return {
        "findings": findings_data,
        "ambiguous": ambiguous_formulas,
        "report_path": report_path,
        "json_path": json_path
    }


# ============================================================================
# CLI Entry Point
# ============================================================================

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Excel Model Auditor")
    parser.add_argument("file_path", help="Path to the Excel file to audit")
    parser.add_argument("--sheets", help="Comma-separated list of sheets to audit", default=None)
    parser.add_argument("--checks", help=f"Comma-separated checks: {','.join(ALL_CHECKS)}", default=None)
    parser.add_argument("--output-format", choices=["xlsx", "md", "html"], default="xlsx")
    parser.add_argument("--output-path", help="Directory for output files", default=None)
    parser.add_argument("--severity", choices=["critical", "warning", "info", "all"], default="all")

    args = parser.parse_args()

    sheets = args.sheets.split(",") if args.sheets else None
    checks = args.checks.split(",") if args.checks else None

    result = run_audit(
        file_path=args.file_path,
        sheets=sheets,
        checks=checks,
        output_format=args.output_format,
        output_path=args.output_path,
        min_severity=args.severity
    )
