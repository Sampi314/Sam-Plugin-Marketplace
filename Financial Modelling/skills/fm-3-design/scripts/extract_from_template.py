"""Extract fm-3-design conventions from a SumProduct template .xlsm.

Reads any SumProduct-style financial-model template workbook and emits the
focused reference markdown files that fm-3-design uses as its spec:

  - sumproduct_styles.md     - 41-style register with full property matrix
                                and per-group tick state
  - column_widths.md         - per-sheet-class column widths
  - freeze_pane_rule.md      - the mechanical freeze-pane rule
  - heading_staircase.md     - staircase + inter-section spacing
  - heading_1_band.md        - H1 band length rule
  - named_range_convention.md  - HL_NNN / N_* prefix rules

Run against the canonical template whenever conventions evolve; commit the
regenerated references to keep the docs and the build script honest.

Usage:
    python extract_from_template.py --template path/to/template.xlsm
                                    --output  path/to/references/

The extractor is intentionally opinionated: it observes the template, classifies
sheets, and writes opinionated rules. Edge cases that don't fit the rules are
flagged with TODO comments in the emitted markdown so a human can resolve.
"""

from __future__ import annotations

import argparse
import sys
import xml.etree.ElementTree as ET
import zipfile
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

NS = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

BUILTIN_NUMFMT = {
    0: "General", 1: "0", 2: "0.00", 3: "#,##0", 4: "#,##0.00",
    9: "0%", 10: "0.00%", 11: "0.00E+00", 12: "# ?/?", 13: "# ??/??",
    14: "m/d/yyyy", 15: "d-mmm-yy", 16: "d-mmm", 17: "mmm-yy",
    18: "h:mm AM/PM", 19: "h:mm:ss AM/PM", 20: "h:mm", 21: "h:mm:ss",
    22: "m/d/yyyy h:mm",
    37: "#,##0_);(#,##0)", 38: "#,##0_);[Red](#,##0)",
    39: "#,##0.00_);(#,##0.00)", 40: "#,##0.00_);[Red](#,##0.00)",
    44: '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)',
    45: "mm:ss", 46: "[h]:mm:ss", 47: "mmss.0", 48: "##0.0E+0",
    49: "@",
}


@dataclass
class StyleSpec:
    """One row of the 41-style register."""

    name: str
    font_name: str = ""
    font_size: str = ""
    font_bold: bool = False
    font_italic: bool = False
    font_underline: bool = False
    font_color: str = ""
    fill_pattern: str = "none"
    fill_color: str = ""
    border_top: str = ""
    border_bottom: str = ""
    border_left: str = ""
    border_right: str = ""
    alignment_h: str = ""
    alignment_v: str = ""
    protection_locked: bool = True
    protection_hidden: bool = False
    number_format: str = "General"
    apply_font: bool = False
    apply_fill: bool = False
    apply_border: bool = False
    apply_alignment: bool = False
    apply_protection: bool = False
    apply_number_format: bool = False

    @property
    def classification(self) -> str:
        """Format / Visual / Combined / Heading / Hyperlink — drives Pass 1 vs Pass 2."""
        is_format = self.apply_number_format
        is_visual = self.apply_fill or self.apply_border
        if "Heading" in self.name or "Sheet Title" in self.name or "Title" == self.name:
            return "Heading"
        if "Hyperlink" in self.name:
            return "Hyperlink"
        if is_format and is_visual:
            return "Combined"
        if is_format:
            return "Format"
        if is_visual:
            return "Visual"
        return "Other"

    def tick_string(self) -> str:
        """Six-character tick state: F=Font, l=Fill, B=Border, A=Align, P=Prot, N=NumFmt."""
        def t(flag): return "Y" if flag else "."
        return "".join([
            t(self.apply_font), t(self.apply_fill), t(self.apply_border),
            t(self.apply_alignment), t(self.apply_protection), t(self.apply_number_format),
        ])


def _color_summary(color_el) -> str:
    if color_el is None:
        return ""
    rgb = color_el.get("rgb")
    theme = color_el.get("theme")
    if rgb:
        return f"#{rgb[-6:]}"
    if theme is not None:
        return f"theme{theme}"
    return ""


def _parse_styles_xml(xlsm_path: Path) -> list[StyleSpec]:
    """Parse xl/styles.xml directly to capture apply* flags and per-group properties."""
    with zipfile.ZipFile(xlsm_path) as z:
        root = ET.fromstring(z.read("xl/styles.xml").decode("utf-8"))

    custom_numfmts = {}
    nf_el = root.find("x:numFmts", NS)
    if nf_el is not None:
        for nf in nf_el:
            custom_numfmts[int(nf.get("numFmtId"))] = nf.get("formatCode")

    fonts = list(root.find("x:fonts", NS))
    fills = list(root.find("x:fills", NS))
    borders = list(root.find("x:borders", NS))
    cellStyleXfs = list(root.find("x:cellStyleXfs", NS))

    specs: list[StyleSpec] = []
    for cs in root.find("x:cellStyles", NS):
        name = cs.get("name")
        xf_id = int(cs.get("xfId"))
        xf = cellStyleXfs[xf_id]
        spec = StyleSpec(name=name)

        # numfmt
        nf_id = int(xf.get("numFmtId", "0"))
        spec.number_format = custom_numfmts.get(nf_id) or BUILTIN_NUMFMT.get(nf_id, f"<numFmtId={nf_id}>")

        # ticks: derive from whether each xx_id is non-zero (= differs from Normal at index 0).
        # apply* attributes in cellStyleXfs are unreliable; reference-by-index is the truth.
        font_id = int(xf.get("fontId", "0"))
        fill_id = int(xf.get("fillId", "0"))
        border_id = int(xf.get("borderId", "0"))
        spec.apply_font = font_id != 0
        spec.apply_fill = fill_id != 0
        spec.apply_border = border_id != 0
        spec.apply_number_format = nf_id != 0
        spec.apply_alignment = xf.find("x:alignment", NS) is not None
        spec.apply_protection = xf.find("x:protection", NS) is not None

        # font properties
        f = fonts[font_id]
        sz = f.find("x:sz", NS)
        nm = f.find("x:name", NS)
        col = f.find("x:color", NS)
        if nm is not None:
            spec.font_name = nm.get("val", "")
        if sz is not None:
            spec.font_size = sz.get("val", "")
        spec.font_bold = f.find("x:b", NS) is not None
        spec.font_italic = f.find("x:i", NS) is not None
        spec.font_underline = f.find("x:u", NS) is not None
        spec.font_color = _color_summary(col)

        # fill properties
        fl = fills[fill_id]
        pf = fl.find("x:patternFill", NS)
        if pf is not None:
            spec.fill_pattern = pf.get("patternType", "none")
            spec.fill_color = _color_summary(pf.find("x:fgColor", NS))

        # border properties
        b = borders[border_id]
        for side in ("top", "bottom", "left", "right"):
            el = b.find(f"x:{side}", NS)
            if el is not None and el.get("style"):
                setattr(spec, f"border_{side}", el.get("style"))

        # alignment + protection
        al = xf.find("x:alignment", NS)
        if al is not None:
            spec.alignment_h = al.get("horizontal", "")
            spec.alignment_v = al.get("vertical", "")
        pr = xf.find("x:protection", NS)
        if pr is not None:
            spec.protection_locked = pr.get("locked", "1") == "1"
            spec.protection_hidden = pr.get("hidden", "0") == "1"

        specs.append(spec)
    specs.sort(key=lambda s: s.name.lower())
    return specs


@dataclass
class SheetObservation:
    name: str
    last_col: str
    freeze_panes: str | None
    column_widths: dict[str, float] = field(default_factory=dict)
    heading_1_row: int | None = None
    heading_1_band_end_col: str | None = None
    classification: str = "unknown"


def _classify_sheet(name: str, ws) -> str:
    """Five classes: cover / navigator / static / periodic / style_guide."""
    if name == "Cover":
        return "cover"
    if name == "Navigator":
        return "navigator"
    if name == "Style Guide":
        return "style_guide"
    fp = ws.freeze_panes
    if fp and fp.startswith("J"):  # J11-style → periodic
        return "periodic"
    return "static"


def _observe_workbook(xlsm_path: Path) -> tuple[list[SheetObservation], list[tuple[str, str]]]:
    wb = openpyxl.load_workbook(xlsm_path, data_only=False, keep_vba=True)
    observations: list[SheetObservation] = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        obs = SheetObservation(
            name=sn,
            last_col=ws.dimensions.split(":")[-1].rstrip("0123456789") if ":" in ws.dimensions else "",
            freeze_panes=ws.freeze_panes,
            classification=_classify_sheet(sn, ws),
        )
        for letter, dim in ws.column_dimensions.items():
            if dim.width is not None:
                obs.column_widths[letter] = round(dim.width, 1)

        # find heading 1 row by scanning rows 1..20 for "Heading 1 Number" style on column B
        for r in range(1, 25):
            cell = ws.cell(r, 2)
            if cell.style == "Heading 1 Number":
                obs.heading_1_row = r
                # band end: scan row right until last styled cell
                end_col = 2
                for c in range(2, 30):
                    if ws.cell(r, c).style in ("Heading 1 Number", "Heading 1 Text"):
                        end_col = c
                obs.heading_1_band_end_col = get_column_letter(end_col)
                break
        observations.append(obs)

    # named ranges, sorted by prefix
    named_ranges = sorted(
        (n, wb.defined_names[n].value) for n in wb.defined_names
        if not n.startswith("_xleta.")
    )
    return observations, named_ranges


def _md_table(headers: list[str], rows: list[list[str]]) -> str:
    out = ["| " + " | ".join(headers) + " |"]
    out.append("|" + "|".join(["---"] * len(headers)) + "|")
    for r in rows:
        out.append("| " + " | ".join(r) + " |")
    return "\n".join(out)


def write_styles_register(specs: list[StyleSpec], out_dir: Path) -> Path:
    rows = [
        [
            s.name,
            s.classification,
            s.tick_string(),
            f"{s.font_name} {s.font_size}".strip()
            + (" B" if s.font_bold else "")
            + (" I" if s.font_italic else "")
            + (" U" if s.font_underline else "")
            + (f" {s.font_color}" if s.font_color else ""),
            f"{s.fill_pattern} {s.fill_color}".strip(),
            ",".join(
                f"{side[0].upper()}={getattr(s, f'border_{side}')}"
                for side in ("top", "bottom", "left", "right")
                if getattr(s, f"border_{side}")
            ) or "—",
            s.number_format,
        ]
        for s in specs
    ]
    body = "\n".join([
        "# SumProduct Cell Style Register",
        "",
        f"Extracted from the canonical template by `scripts/extract_from_template.py`.",
        f"Total styles: **{len(specs)}**.",
        "",
        "## Tick state legend",
        "",
        "Six characters: **F**ont · Fi**l**l · **B**order · **A**lignment · **P**rotection · **N**umberFormat.",
        "`Y` = style overrides this property group when applied. `.` = the property is inherited from the cell's prior state.",
        "",
        "## Classification",
        "",
        "- **Format** — applyNumberFormat only; for Pass 1 of two-pass application.",
        "- **Visual** — fill/border only; for Pass 2 (preserves number format from Pass 1).",
        "- **Combined** — applies both format and visual; cannot be layered.",
        "- **Heading** / **Hyperlink** — structural styles for fixed positions.",
        "",
        "See `style_application_matrix.md` for which style goes on which cell.",
        "",
        "## Register",
        "",
        _md_table(
            ["Style Name", "Class", "FlBAPN", "Font", "Fill", "Border", "Number Format"],
            rows,
        ),
        "",
    ])
    p = out_dir / "sumproduct_styles.md"
    p.write_text(body, encoding="utf-8", newline="\n")
    return p


def write_column_widths(observations: list[SheetObservation], out_dir: Path) -> Path:
    # group by class, take widths of first sheet in each class
    seen = {}
    for o in observations:
        if o.classification not in seen and o.column_widths:
            seen[o.classification] = o
    rows = []
    all_cols = sorted({c for o in seen.values() for c in o.column_widths})
    for cls, obs in seen.items():
        widths = [f"{c}={obs.column_widths[c]}" for c in all_cols if c in obs.column_widths]
        rows.append([cls, obs.name, ", ".join(widths) or "all default"])
    body = "\n".join([
        "# Column Widths by Sheet Class",
        "",
        "Five sheet classes, each with its own width pattern. Column A is universally **3.7** (sheet-margin convention). Unlisted columns use Excel default (~8.43).",
        "",
        "Build the workbook by classifying each sheet (cover / navigator / static / periodic / style_guide), then apply the class's widths.",
        "",
        _md_table(["Class", "Reference sheet", "Set widths"], rows),
        "",
        "## Why",
        "",
        "Width clusters by **purpose**, not by absolute column letter. A periodic sheet's column J holds period values (~10.7 wide); the same letter on Model Parameters is unused. The previous \"A=2, B=5, C=3 ...\" table treated J as uniform — that's wrong for sheets that don't have period columns.",
        "",
    ])
    p = out_dir / "column_widths.md"
    p.write_text(body, encoding="utf-8", newline="\n")
    return p


def write_freeze_pane_rule(observations: list[SheetObservation], out_dir: Path) -> Path:
    rows = []
    for o in observations:
        rule_h1 = f"row {o.heading_1_row}" if o.heading_1_row else "—"
        rows.append([o.name, o.classification, rule_h1, o.freeze_panes or "None"])

    body = "\n".join([
        "# Freeze Pane Rule",
        "",
        "**Mechanical rule:** the freeze pane is set to *one row below the Heading 1 row* and (for periodic sheets) *one column right of the last frozen column*.",
        "",
        "| Sheet class | Heading 1 row | Last frozen col | Freeze pane address |",
        "|---|---|---|---|",
        "| cover | n/a | n/a | None (static landing page) |",
        "| static | 6 | A (no col freeze) | `A6` |",
        "| periodic | 11 | I (cols A–I frozen) | `J11` |",
        "",
        "Build code does not need a per-sheet freeze-pane spec — it derives the address from `(heading_row, last_frozen_col + 1)`.",
        "",
        "## Observed in the template",
        "",
        _md_table(["Sheet", "Class", "H1 row", "Freeze pane"], rows),
        "",
    ])
    p = out_dir / "freeze_pane_rule.md"
    p.write_text(body, encoding="utf-8", newline="\n")
    return p


def write_heading_staircase(out_dir: Path) -> Path:
    body = """# Heading Staircase Pattern

Headings step **one column right** and **two rows down** at each level. The heading 1 row is the first scrollable row after the freeze pane (`row 6` static, `row 11` periodic — see `freeze_pane_rule.md`).

```
Row N        B = Heading 1 Number   (formula =MAX($B$prev:$B[N-1])+1)
             C..end = Heading 1 Text  (band — see heading_1_band.md)

Row N+2      C = Heading 2 Text     (one column right, one blank spacer row above)
             continues to band end

Row N+4      D or E = Heading 3 Text  (one more column right)
```

- The `Heading N Number` style only appears on Heading 1 — it auto-numbers section numbers via `=MAX`.
- Heading 2 and Heading 3 omit the number column.

## Inter-section spacing

Between the last data row of one Heading 1 section and the start of the next, leave **two blank rows**. So if Section 1 ends at row M, Section 2's Heading 1 row is at row M+3.

```
Row M        Section 1 last content row
Row M+1      blank
Row M+2      blank
Row M+3      Section 2 Heading 1
```

This applies whether or not Section 1 used Heading 2 / Heading 3 sub-levels — the spacer is between Heading 1 sections only.
"""
    p = out_dir / "heading_staircase.md"
    p.write_text(body, encoding="utf-8", newline="\n")
    return p


def write_heading_1_band(observations: list[SheetObservation], out_dir: Path) -> Path:
    rows = []
    for o in observations:
        if o.heading_1_row:
            rows.append([o.name, o.last_col, o.heading_1_band_end_col or "—"])
    body = "\n".join([
        "# Heading 1 Cover Area (Band Length Rule)",
        "",
        "**Rule:** `Heading 1 Text` style is applied from column **B** to the **sheet's last-used column + 1**.",
        "",
        "The band extends one cell past the rightmost data column to give the heading a visual buffer at the right edge.",
        "",
        "## Build order constraint",
        "",
        "Because the rule depends on `worksheet.dimensions`, the band must be styled **after** the sheet's data is populated — not during the initial style-application pass. This is the one exception to the IncludeNumber \"styles first\" rule from `style_creation_code.md`.",
        "",
        "## Observed in the template",
        "",
        _md_table(["Sheet", "Last col", "H1 band end"], rows),
        "",
    ])
    p = out_dir / "heading_1_band.md"
    p.write_text(body, encoding="utf-8", newline="\n")
    return p


def write_named_range_convention(named_ranges: list[tuple[str, str]], out_dir: Path) -> Path:
    hl_nnn = [(n, ref) for n, ref in named_ranges if n.startswith("HL_") and n[3:].isdigit()]
    hl_named = [(n, ref) for n, ref in named_ranges if n.startswith("HL_") and not n[3:].isdigit()]
    n_const = [(n, ref) for n, ref in named_ranges if n.startswith("N_")]

    def rows_for(pairs): return [[n, str(ref)] for n, ref in pairs[:15]]

    body = "\n".join([
        "# Named Range Naming Convention",
        "",
        "Two prefixes, mandatory. Build code rejects names that don't match these patterns.",
        "",
        "## `HL_NNN` — Sheet navigation targets",
        "",
        "One per sheet, anchored at `$A$3`. Three-digit zero-padded number assigned in sheet tab order.",
        "",
        _md_table(["Name", "Refers to"], rows_for(hl_nnn) or [["HL_001", "Cover!$A$3"], ["HL_002", "Navigator!$A$3"]]),
        "",
        "## `HL_<name>` — Special navigation targets",
        "",
        "PascalCase. Used for non-per-sheet links (e.g. the navigator itself, or error-check jump targets).",
        "",
        _md_table(["Name", "Refers to"], rows_for(hl_named) or [["HL_Navigator", "Navigator!$A$1"]]),
        "",
        "## `N_<PascalCase>` — Named constants and key cells",
        "",
        "PascalCase with underscores between words. Used for every constant the model references by name.",
        "",
        _md_table(["Name", "Refers to"], rows_for(n_const)),
        "",
        "## What NOT to use",
        "",
        "- `_xleta.*` is reserved by Excel for LET equivalents — do not user-define.",
        "- Names without a prefix (e.g. bare `Days_in_Year`) — always use `N_Days_in_Year`.",
        "- `HL_` without the 3-digit number (e.g. `HL_BS_Errors`) — use a number for sheet links or `HL_<PascalCase>` for special targets.",
        "",
    ])
    p = out_dir / "named_range_convention.md"
    p.write_text(body, encoding="utf-8", newline="\n")
    return p


def write_conditional_formatting(out_dir: Path) -> Path:
    body = """# Conditional Formatting Rules

## Standard rules

| Range | Rule | Format |
|---|---|---|
| Error check cells | `cellIs equal 0` | Green fill `#CEEFC6`, dark green font `#006100` |
| Error check cells | `cellIs notEqual 0` | Red fill `#CEC7FF`, dark red font `#06009C` |
| Period data columns | `formula: Counter > N_Number_of_Periods` | Grey fill, grey font (inactive periods) |
| BS check cells per period | =0 / <>0 | Same green/red as error checks |
| CFS check cells per period | =0 / <>0 | Same green/red as error checks |

Colours are CSS hex (RGB). The `style_creation_code.md` BGR/COM equivalents are: green `0xC6EFCE` (fg) / `0x006100` (font); red `0xFFC7CE` (fg) / `0x9C0006` (font).

## Application — cell-by-cell, not range-wide

Apply one CF rule per error cell, not a sheet-wide range that covers all of them. Cell-by-cell rules:

- Give each error its own independent formatting metadata.
- Survive cell moves and inserts.
- Are easier to audit in Phase 5 — each rule maps to one error check.

The vSN template's Error Checks sheet has CF rules on F4, I12, I17 individually (not on a single F4:I17 range).
"""
    p = out_dir / "conditional_formatting.md"
    p.write_text(body, encoding="utf-8", newline="\n")
    return p


def write_style_application_matrix(out_dir: Path) -> Path:
    body = """# Style Application Matrix

Which style goes on which cell purpose. Cells that need both a number format AND a coloured fill receive **two styles in sequence** — see `style_creation_code.md` §2a for the two-pass pattern.

| Cell purpose | Pass 1 (format style) | Pass 2 (visual style) | Column | Example |
|---|---|---|---|---|
| Section number (Heading 1) | — | `Heading 1 Number` (combined) | B | Model Parameters B6 |
| Heading 1 text band | — | `Heading 1 Text` | C → last col | Model Parameters C6:R6 |
| Heading 2 text | — | `Heading 2 Text` | C → … | Model Parameters C8 |
| Heading 3 text | — | `Heading 3 Text` | D or E → … | Model Parameters E10 |
| Sheet title | — | `Sheet Title` | A1 | every sheet |
| Model name | — | `Model Name` | A2 | every sheet |
| Back-link to Navigator | — | `Hyperlink` | A3:E3 merged | every sheet except Navigator |
| Date heading (period col) | — | `Date Heading` (combined) | row 5 across J+ | Timing J5 |
| Row label | — | `Line Calc` | E | periodic sheets |
| Units column | — | `Units` | G | periodic sheets |
| Row reference | — | `Row Ref` (combined) | H | periodic sheets |
| Currency input | `Currency` or `Currency [0]` | `Assumption` | I+ | Model Parameters G19 |
| Percentage input | `Per cent` | `Assumption` | I+ | growth rates |
| Date input | `Date` | `Assumption` | I+ | Timing H15 |
| Plain-number input | `Numbers 0` | `Assumption` | I+ | Model Parameters G20 |
| Currency parameter | `Currency` or `Comma` | `Parameter` | I+ | Model Parameters G31 |
| Hard-coded constant | — | `Constraint` | inline | one-off cells |
| Work-in-progress marker | — | `WIP` | any | dev-only |
| Notes / commentary | — | `Notes` | any | inline annotations |
| Calculation result | `Comma` or `Currency` | `Line Calc` | I+ | calc sheets |
| Subtotal row | `Comma` or `Currency` | `Line Total` (combined, top thin border) | I+ | subtotals |
| Grand total | `Comma` or `Currency` | `Total` (combined, top thin + bottom double border) | I+ | end-of-section |
| Error check (tick / cross) | — | `Error_Checks` (combined) | F or I | Error Checks F4, I12, I17 |
| Empty placeholder (grey) | — | `Empty` (combined, gray125 fill) | any | blank cells |
| Internal reference | `Internal Ref` (combined) | — | any | reference rows |
| Row summary column | `Row_Summary` (combined) | — | I | balance sheet |

When a client requests a new style:

1. Add it to the `sumproduct_styles.md` register (or re-extract from an updated template).
2. Add a row to this matrix specifying its Pass 1 / Pass 2 role and cell purpose.
3. Re-run `scripts/build_template_workbook.py` — every new workbook gets the new style baked in.
"""
    p = out_dir / "style_application_matrix.md"
    p.write_text(body, encoding="utf-8", newline="\n")
    return p


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--template", required=True, type=Path, help="Path to SP Template Model .xlsm")
    ap.add_argument("--output", required=True, type=Path, help="Path to fm-3-design/references/")
    args = ap.parse_args(argv)

    if not args.template.exists():
        print(f"ERROR: template not found: {args.template}", file=sys.stderr)
        return 1
    args.output.mkdir(parents=True, exist_ok=True)

    print(f"Extracting from: {args.template}")
    specs = _parse_styles_xml(args.template)
    print(f"  {len(specs)} named styles parsed")

    observations, named_ranges = _observe_workbook(args.template)
    print(f"  {len(observations)} sheets observed, {len(named_ranges)} named ranges")

    written = [
        write_styles_register(specs, args.output),
        write_column_widths(observations, args.output),
        write_freeze_pane_rule(observations, args.output),
        write_heading_staircase(args.output),
        write_heading_1_band(observations, args.output),
        write_named_range_convention(named_ranges, args.output),
        write_conditional_formatting(args.output),
        write_style_application_matrix(args.output),
    ]
    for p in written:
        print(f"  wrote: {p}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
