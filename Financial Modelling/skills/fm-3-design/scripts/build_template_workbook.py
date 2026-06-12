"""Build a clean SumProduct starter workbook with every fm-3-design convention.

Produces a .xlsx that already carries:
  - All 41 named styles from sumproduct_styles.md
  - A Style Guide sheet documenting them
  - The standard skeleton sheets (Cover, Navigator, Style Guide, Model
    Parameters, Timing, Error Checks, Change Log, Timing Template,
    No Timing Template) — each with A1 sheet-name formula, A2 model-name
    formula, A3:E3 merged Navigator back-link
  - Standard column widths per sheet class
  - Freeze panes per the mechanical rule (A6 for static, J11 for periodic,
    none for Cover)
  - Heading 1 rows with auto-numbering formula and the full Heading 1 Text
    band (B to last-used col + 1)
  - All HL_NNN / HL_Navigator / N_* named ranges per the standard list

The script is idempotent: re-running it on the same workbook adds nothing
new (styles caught with try/except, sheets created only if absent).

fm-4-build consumes the produced workbook and adds model-specific content.
It does NOT re-derive conventions — they're already baked in here.

Usage:
    python build_template_workbook.py --output new_model.xlsx
    python build_template_workbook.py --output new_model.xlsx \
        --sheets Cover Navigator "Style Guide" "Model Parameters" Timing "Error Checks"
"""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, NamedStyle, PatternFill, Protection, Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.hyperlink import Hyperlink

# --------------------------------------------------------------------------
# Standard inputs (mirror the references)
# --------------------------------------------------------------------------

DEFAULT_SHEETS = [
    "Cover", "Navigator", "Style Guide", "Model Parameters", "Timing",
    "Error Checks", "Change Log", "Timing Template", "No Timing Template",
]

# Sheet classification — drives freeze pane and column widths
SHEET_CLASS = {
    "Cover": "cover",
    "Navigator": "navigator",
    "Style Guide": "style_guide",
    "Timing": "periodic",
    "Timing Template": "periodic",
}
# Anything not listed -> "static"

# Column widths by class — only explicit (others use Excel default ~8.43)
COLUMN_WIDTHS = {
    "cover":      {"A": 3.7, "C": 3.7},
    "navigator":  {"A": 3.7, "F": 17.7},
    "static":     {"A": 3.7, "F": 11.1, "G": 27.0, "H": 21.6,
                   "I": 24.4, "J": 30.3, "K": 26.7},
    "periodic":   {"A": 3.7, "G": 22.1, "H": 10.7, "J": 10.7,
                   "K": 10.7, "L": 10.7, "M": 10.7, "N": 10.7, "O": 10.7},
    "style_guide":{"A": 3.7, "F": 9.1, "H": 1.7, "I": 17.3, "J": 1.7,
                   "K": 23.4, "N": 1.7, "P": 9.1},
}

# Freeze pane by class
FREEZE_PANE = {
    "cover":       None,
    "navigator":   "A6",
    "style_guide": "A6",
    "static":      "A6",
    "periodic":    "J11",
}

# Heading 1 row by class
HEADING_1_ROW = {
    "cover": None,
    "navigator": 6,
    "style_guide": 6,
    "static": 6,
    "periodic": 11,
}

# Sheet title formula — extracts sheet name from file path
SHEET_TITLE_FORMULA = (
    '=IF(ISERROR(RIGHT(CELL("filename",A2),'
    'LEN(CELL("filename",A2))-FIND("]",CELL("filename",A2)))),"",'
    'RIGHT(CELL("filename",A2),'
    'LEN(CELL("filename",A2))-FIND("]",CELL("filename",A2))))'
)

# Standard named constants on Model Parameters / Timing
# (name, sheet, cell, default_value, format_style, visual_style)
NAMED_CONSTANTS = [
    # Model Parameters — General block
    ("N_Model_Name",            "Model Parameters", "G11", "[Model Name]",           None,         "Assumption"),
    ("N_Client_Name",           "Model Parameters", "G12", "[Client Name]",          None,         "Assumption"),
    ("N_Days_in_Year",          "Model Parameters", "G19", 365,                      "Numbers 0",  "Assumption"),
    ("N_Months_in_Month",       "Model Parameters", "G20", 1,                        "Numbers 0",  "Assumption"),
    ("N_Months_in_Qtr",         "Model Parameters", "G21", 3,                        "Numbers 0",  "Assumption"),
    ("N_Months_in_Half_Yr",     "Model Parameters", "G22", 6,                        "Numbers 0",  "Assumption"),
    ("N_Months_in_Year",        "Model Parameters", "G23", 12,                       "Numbers 0",  "Assumption"),
    ("N_Quarters_in_Year",      "Model Parameters", "G24", 4,                        "Numbers 0",  "Assumption"),
    ("N_Rounding_Accuracy",     "Model Parameters", "G26", 0.0001,                   "Per cent",   "Parameter"),
    ("N_Very_Large_Number",     "Model Parameters", "G28", 1e99,                     None,         "Parameter"),
    ("N_Very_Small_Number",     "Model Parameters", "G29", 1e-99,                    None,         "Parameter"),
    ("N_Thousand",              "Model Parameters", "G31", 1000,                     "Numbers 0",  "Parameter"),
    # Timing
    ("N_Model_Start_Date",      "Timing",           "H15", "=DATE(2025,1,1)",        "Date",       "Assumption"),
    ("N_Periodicity",           "Timing",           "H17", 3,                        "Numbers 0",  "Assumption"),
    ("N_Example_Reporting_Month","Timing",          "H19", 6,                        "Numbers 0",  "Assumption"),
    ("N_Reporting_Month_Factor","Timing",           "H21", "=H19/H17",               "Numbers 0",  None),
    # Error Checks
    ("N_Overall_Error_Check",   "Error Checks",     "I17", "=SUMIF(F4:F12,\"<>0\",F4:F12)", None,  None),
]

# --------------------------------------------------------------------------
# Style parsing — pull specs out of sumproduct_styles.md so the script and
# the reference stay in sync
# --------------------------------------------------------------------------

THEME_COLORS = {
    # Approximations of Office theme colours (theme1 = light1/dark1 mapping)
    "theme0": "000000", "theme1": "FFFFFF", "theme2": "44546A",
    "theme3": "E7E6E6", "theme4": "4472C4", "theme5": "ED7D31",
    "theme6": "A5A5A5", "theme7": "FFC000", "theme8": "5B9BD5",
    "theme9": "70AD47",
}


@dataclass
class StyleSpec:
    name: str
    classification: str
    ticks: str  # FlBAPN string
    font_str: str
    fill_str: str
    border_str: str
    number_format: str


def parse_style_register(register_path: Path) -> list[StyleSpec]:
    """Parse the markdown table in sumproduct_styles.md into StyleSpec list."""
    text = register_path.read_text(encoding="utf-8")
    # Find the table after "## Register"
    m = re.search(r"## Register\s*\n\n(\|.*?)(?:\n\n|\Z)", text, re.DOTALL)
    if not m:
        raise RuntimeError(f"Could not find Register table in {register_path}")
    table = m.group(1)
    rows = [r for r in table.strip().split("\n") if r.startswith("|")]
    if len(rows) < 3:
        raise RuntimeError(f"Register table too short in {register_path}")
    # Skip header (row 0) + separator (row 1)
    specs: list[StyleSpec] = []
    for r in rows[2:]:
        cells = [c.strip() for c in r.strip("|").split("|")]
        if len(cells) < 7:
            continue
        specs.append(StyleSpec(
            name=cells[0],
            classification=cells[1],
            ticks=cells[2],
            font_str=cells[3],
            fill_str=cells[4],
            border_str=cells[5],
            number_format=cells[6],
        ))
    return specs


def _resolve_color(token: str) -> str:
    """Convert '#FF0000' or 'theme8' or '' to a 6-char hex RGB."""
    token = token.strip()
    if not token:
        return "000000"
    if token.startswith("#"):
        return token[1:].rjust(6, "0").upper()
    if token in THEME_COLORS:
        return THEME_COLORS[token]
    return "000000"


def _build_font(spec: StyleSpec) -> Font:
    """Parse 'Arial 9 B I U theme1' into Font."""
    parts = spec.font_str.split()
    name = parts[0] if parts else "Calibri"
    size = 11
    bold = italic = underline = False
    color = "000000"
    i = 1
    while i < len(parts):
        p = parts[i]
        if p.isdigit():
            size = int(p)
        elif p == "B":
            bold = True
        elif p == "I":
            italic = True
        elif p == "U":
            underline = True
        elif p.startswith("#") or p.startswith("theme"):
            color = _resolve_color(p)
        i += 1
    return Font(
        name=name, size=size, bold=bold, italic=italic,
        underline="single" if underline else None,
        color=color,
    )


def _build_fill(spec: StyleSpec) -> PatternFill:
    """Parse 'solid #FFFF99' or 'gray125 theme8' or 'none' into PatternFill."""
    s = spec.fill_str.strip()
    if not s or s == "none":
        return PatternFill(fill_type=None)
    parts = s.split(None, 1)
    pattern = parts[0]
    color = _resolve_color(parts[1]) if len(parts) > 1 else "FFFFFF"
    return PatternFill(fill_type=pattern, fgColor=color, bgColor=color)


def _build_border(spec: StyleSpec) -> Border:
    """Parse 'T=thin,B=double,L=thin,R=thin' into Border."""
    if not spec.border_str or spec.border_str == "—":
        return Border()
    sides_map = {"T": "top", "B": "bottom", "L": "left", "R": "right"}
    kwargs = {}
    for part in spec.border_str.split(","):
        part = part.strip()
        if "=" not in part:
            continue
        side_key, style = part.split("=", 1)
        side_name = sides_map.get(side_key.strip())
        if side_name:
            kwargs[side_name] = Side(style=style.strip(), color="000000")
    return Border(**kwargs)


def specs_to_named_styles(specs: list[StyleSpec]) -> list[NamedStyle]:
    out = []
    for s in specs:
        # Only set property groups whose tick is Y in the FlBAPN string
        font = _build_font(s) if s.ticks[0] == "Y" else None
        fill = _build_fill(s) if s.ticks[1] == "Y" else None
        border = _build_border(s) if s.ticks[2] == "Y" else None
        nf = s.number_format if s.ticks[5] == "Y" else "General"
        ns = NamedStyle(name=s.name, number_format=nf)
        if font: ns.font = font
        if fill: ns.fill = fill
        if border: ns.border = border
        if s.ticks[3] == "Y":
            ns.alignment = Alignment(horizontal="general", vertical="bottom")
        out.append(ns)
    return out


# --------------------------------------------------------------------------
# Build phases
# --------------------------------------------------------------------------

def phase_a_create_styles(wb: Workbook, specs: list[StyleSpec]) -> int:
    """Add all named styles to wb. Idempotent — skips duplicates."""
    added = 0
    existing = set(wb.named_styles)  # list[str] of names
    for ns in specs_to_named_styles(specs):
        if ns.name in existing:
            continue
        try:
            wb.add_named_style(ns)
            existing.add(ns.name)
            added += 1
        except (ValueError, KeyError):
            pass  # already present in alternative form
    return added


def phase_b_skeleton_sheets(wb: Workbook, sheet_names: list[str]) -> dict[str, str]:
    """Create skeleton sheets with A1/A2/A3, widths, freeze panes."""
    classes = {}
    # Drop the default "Sheet" tab if it's empty
    if "Sheet" in wb.sheetnames and len(wb["Sheet"]._cells) == 0 and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    for name in sheet_names:
        cls = SHEET_CLASS.get(name, "static")
        classes[name] = cls
        if name in wb.sheetnames:
            ws = wb[name]
        else:
            ws = wb.create_sheet(name)

        # A1 — Sheet Title
        ws["A1"] = SHEET_TITLE_FORMULA
        ws["A1"].style = "Sheet Title"
        # A2 — Model Name
        ws["A2"] = "=N_Model_Name"
        ws["A2"].style = "Model Name"
        # A3 — Navigator back-link (everywhere except Navigator itself)
        if name != "Navigator":
            ws["A3"] = "Navigator"
            ws["A3"].style = "Hyperlink"
            ws["A3"].hyperlink = Hyperlink(ref="A3", location="HL_Navigator",
                                           tooltip="Go to Navigator (Table of Contents)",
                                           display="Navigator")
            if "A3:E3" not in {str(r) for r in ws.merged_cells.ranges}:
                ws.merge_cells("A3:E3")

        # Column widths
        for letter, w in COLUMN_WIDTHS[cls].items():
            ws.column_dimensions[letter].width = w

        # Freeze pane
        fp = FREEZE_PANE[cls]
        if fp:
            ws.freeze_panes = fp

    # Drop "Sheet" if it survived because the workbook came in empty
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        ws_sheet = wb["Sheet"]
        if all(c.value is None for r in ws_sheet.iter_rows() for c in r):
            del wb["Sheet"]

    return classes


def phase_c_named_ranges(wb: Workbook, sheet_names: list[str]) -> int:
    """Register HL_NNN per sheet + HL_Navigator + N_* constants."""
    added = 0
    # HL_NNN — per sheet, anchored at $A$3
    for i, name in enumerate(sheet_names, start=1):
        hl_name = f"HL_{i:03d}"
        ref = f"'{name}'!$A$3"
        if hl_name not in wb.defined_names:
            wb.defined_names[hl_name] = DefinedName(name=hl_name, attr_text=ref)
            added += 1
    # HL_Navigator
    if "HL_Navigator" not in wb.defined_names and "Navigator" in sheet_names:
        wb.defined_names["HL_Navigator"] = DefinedName(
            name="HL_Navigator", attr_text="'Navigator'!$A$1")
        added += 1
    # N_* constants
    for n_name, sheet, cell, value, fmt_style, vis_style in NAMED_CONSTANTS:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        ws[cell] = value
        if fmt_style:
            ws[cell].style = fmt_style
        if vis_style:
            ws[cell].style = vis_style  # Pass 2 — overlays format
        if n_name not in wb.defined_names:
            ref = f"'{sheet}'!${cell[0]}${cell[1:]}"
            wb.defined_names[n_name] = DefinedName(name=n_name, attr_text=ref)
            added += 1
    return added


def phase_d_heading_staircase(wb: Workbook, classes: dict[str, str]) -> int:
    """Apply Heading 1 row to working sheets — B Number, C..last+1 Text band."""
    applied = 0
    for sheet_name, cls in classes.items():
        h1_row = HEADING_1_ROW[cls]
        if h1_row is None:
            continue
        if sheet_name in ("Navigator", "Style Guide"):
            continue  # built separately
        ws = wb[sheet_name]
        # Section number in column B
        ws.cell(h1_row, 2).value = f"=MAX($B$5:$B{h1_row-1})+1"
        ws.cell(h1_row, 2).style = "Heading 1 Number"
        # Heading 1 Text band: from C to (last data col + 1), or to col J as minimum
        last_col = 10  # default to J
        # If sheet has data already, extend band
        if ws.max_column and ws.max_column > last_col:
            last_col = ws.max_column + 1
        for c in range(3, last_col + 1):
            ws.cell(h1_row, c).style = "Heading 1 Text"
        # The first section starts empty — user will fill the label
        ws.cell(h1_row, 3).value = "[Section name]"
        applied += 1
    return applied


def phase_e_style_guide_sheet(wb: Workbook) -> bool:
    """Build the Style Guide sheet — 3 sections documenting every style."""
    if "Style Guide" not in wb.sheetnames:
        return False
    ws = wb["Style Guide"]

    # Section 1 — Formatting of Headers / Dividers
    ws.cell(6, 2).value = "=MAX($B$5:$B5)+1"
    ws.cell(6, 2).style = "Heading 1 Number"
    ws.cell(6, 3).value = "Formatting of Headers / Dividers"
    ws.cell(6, 3).style = "Heading 1 Text"
    for c in range(4, 14):
        ws.cell(6, c).style = "Heading 1 Text"

    _table_header(ws, 8)
    headers = [
        (10, "Sheet Title"), (11, "Model Name"), (13, "Header 1"),
        (14, "Header 2"), (15, "Header 3"), (16, "Header 4"),
        (18, "Notes"), (20, "Table Heading"),
    ]
    style_map_1 = {
        "Sheet Title": "Sheet Title", "Model Name": "Model Name",
        "Header 1": "Heading 1 Text", "Header 2": "Heading 2 Text",
        "Header 3": "Heading 3 Text", "Header 4": "Heading 4",
        "Notes": "Notes", "Table Heading": "Table_Heading",
    }
    for row, desc in headers:
        ws.cell(row, 3).value = desc
        ws.cell(row, 9).value = desc
        ws.cell(row, 9).style = style_map_1[desc]
        ws.cell(row, 11).value = style_map_1[desc]

    # Section 2 — Individual Cell Styles
    ws.cell(23, 2).value = "=MAX($B$5:$B22)+1"
    ws.cell(23, 2).style = "Heading 1 Number"
    ws.cell(23, 3).value = "Individual Cell Styles"
    ws.cell(23, 3).style = "Heading 1 Text"
    for c in range(4, 14):
        ws.cell(23, c).style = "Heading 1 Text"

    _table_header(ws, 25)
    individual = [
        (27, "Assumption", "Assumption", 1234),
        (29, "Constraint", "Constraint", "Fixed"),
        (31, "Empty", "Empty", ""),
        (33, "Error Check", "Error_Checks", 0),
        (35, "Hyperlink", "Hyperlink", "Hyperlink"),
        (37, "Internal Ref", "Internal Ref", "='Error Checks'!E12"),
        (39, "Line Calculation", "Line Calc", 77),
        (41, "Line Total", "Line Total", "=I39"),
        (43, "Parameter", "Parameter", 365),
        (45, "Range Name Description", "Range Name Description", "Not_Named"),
        (47, "Row Reference", "Row Ref", "=ROW(C47)"),
        (49, "Row Summary", "Row_Summary", "=I41"),
        (51, "Units", "Units", "A$"),
        (53, "WIP", "WIP", "TODO"),
    ]
    for row, desc, style, example in individual:
        ws.cell(row, 3).value = desc
        ws.cell(row, 9).value = example
        ws.cell(row, 9).style = style
        ws.cell(row, 11).value = style

    # Section 3 — Numerical Styles
    ws.cell(56, 2).value = "=MAX($B$5:$B55)+1"
    ws.cell(56, 2).style = "Heading 1 Number"
    ws.cell(56, 3).value = "Numerical Styles"
    ws.cell(56, 3).style = "Heading 1 Text"
    for c in range(4, 14):
        ws.cell(56, c).style = "Heading 1 Text"

    _table_header(ws, 58)
    numerical = [
        (60, "Comma", "Comma", 123456.789),
        (62, "Comma [0]", "Comma [0]", -123456.789),
        (64, "Currency", "Currency", 123456.789),
        (66, "Currency [0]", "Currency [0]", 123456.789),
        (68, "Date", "Date", "=TODAY()"),
        (70, "Date Heading", "Date Heading", "=TODAY()"),
        (72, "Numbers 0", "Numbers 0", -123456.789),
        (74, "Percent", "Per cent", 0.5),
    ]
    for row, desc, style, example in numerical:
        ws.cell(row, 3).value = desc
        ws.cell(row, 9).value = example
        ws.cell(row, 9).style = style
        ws.cell(row, 11).value = style

    # Error check label at top
    ws.cell(4, 5).value = "Error Checks:"
    ws.cell(4, 9).value = "=N_Overall_Error_Check"

    return True


def _table_header(ws, row: int):
    """Standard 3-column table header: Description / Display / Style Name."""
    ws.cell(row, 3).value = "Description"
    ws.cell(row, 9).value = "Display"
    ws.cell(row, 11).value = "Style Name"
    for c in (3, 9, 11):
        ws.cell(row, c).style = "Table_Heading"


# --------------------------------------------------------------------------
# Verification
# --------------------------------------------------------------------------

def phase_f_verify(wb_path: Path, expected_styles: int, expected_sheets: list[str]) -> bool:
    """Re-open the produced workbook and verify the conventions stuck."""
    wb = openpyxl.load_workbook(wb_path)
    style_count = len(wb.named_styles)
    if style_count < expected_styles - 3:  # tolerate a few Excel-managed duplicates
        print(f"  WARN: expected ~{expected_styles} styles, got {style_count}")
    for sn in expected_sheets:
        if sn not in wb.sheetnames:
            print(f"  FAIL: sheet '{sn}' missing")
            return False
    if "HL_Navigator" not in wb.defined_names:
        print("  FAIL: HL_Navigator named range missing")
        return False
    print(f"  OK: {style_count} styles, {len(wb.sheetnames)} sheets, "
          f"{len(list(wb.defined_names))} named ranges")
    return True


# --------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------

def main(argv: list[str] | None = None) -> int:
    here = Path(__file__).resolve().parent
    default_register = here.parent / "references" / "sumproduct_styles.md"

    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--output", required=True, type=Path,
                    help="Output .xlsx path (will be created or overwritten)")
    ap.add_argument("--sheets", nargs="*", default=DEFAULT_SHEETS,
                    help="Sheet names to include (default: standard 9)")
    ap.add_argument("--register", type=Path, default=default_register,
                    help=f"Style register markdown (default: {default_register})")
    args = ap.parse_args(argv)

    if not args.register.exists():
        print(f"ERROR: style register not found: {args.register}", file=sys.stderr)
        return 1

    print(f"Building: {args.output}")
    print(f"  Sheets: {args.sheets}")

    specs = parse_style_register(args.register)
    print(f"  Parsed {len(specs)} style specs from register")

    # If output exists, open and update (idempotent re-run); else create new
    if args.output.exists():
        wb = openpyxl.load_workbook(args.output)
        print(f"  Updating existing workbook")
    else:
        wb = Workbook()
        # delete the default "Sheet"
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    print("Phase A — create styles")
    added = phase_a_create_styles(wb, specs)
    print(f"  added {added} new styles ({len(wb.named_styles)} total)")

    print("Phase B — skeleton sheets")
    classes = phase_b_skeleton_sheets(wb, args.sheets)
    print(f"  built {len(classes)} sheets")

    print("Phase C — named ranges")
    n_added = phase_c_named_ranges(wb, args.sheets)
    print(f"  added {n_added} named ranges")

    print("Phase D — heading staircase")
    h_applied = phase_d_heading_staircase(wb, classes)
    print(f"  applied H1 row to {h_applied} working sheets")

    print("Phase E — Style Guide sheet content")
    if phase_e_style_guide_sheet(wb):
        print("  Style Guide sheet built")

    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(f"Saved: {args.output}")

    print("Phase F — verify")
    ok = phase_f_verify(args.output, expected_styles=len(specs),
                        expected_sheets=args.sheets)
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
