"""extract_workbook.py — one openpyxl pass dumps a complete JSON model of a workbook.

Usage:
    python extract_workbook.py <path-to.xlsx|.xlsm> [--out extract.json]
                               [--sheets "Sheet A,Sheet B"] [--digest digest.md]
                               [--max-cells 200000]

Run FIRST in any audit. Every excel-* auditor rule script consumes the JSON
this produces — extract once, audit many times. Works with Excel closed; no
COM required (Power Query / VBA need excel_automation.py instead).

Top-level JSON keys (full schema: ../references/extraction_guide.md):
  meta, cells, r1c1_rows, r1c1_cols, errors, named_ranges, validations,
  conditional_formatting, hyperlinks, text_inventory, dependencies, styles, charts
"""

from __future__ import annotations

import argparse
import datetime as _dt
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from audit_lib import a1_to_r1c1, col_letter, expand_range  # noqa: E402

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    sys.exit("openpyxl is required: pip install openpyxl")

ERROR_LITERALS = {
    "#REF!", "#VALUE!", "#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!",
    "#SPILL!", "#CALC!", "#GETTING_DATA",
}

BUILTIN_STYLES = {
    "Normal", "Comma", "Comma [0]", "Currency", "Currency [0]", "Percent",
    "Hyperlink", "Followed Hyperlink", "Note", "Warning Text", "Title",
    "Heading 1", "Heading 2", "Heading 3", "Heading 4", "Input", "Output",
    "Calculation", "Check Cell", "Explanatory Text", "Linked Cell",
    "Total", "Good", "Bad", "Neutral", "20 % - Accent1", "40 % - Accent1",
}

SHEET_REF_RE = re.compile(r"(?:'([^'\[\]]+)'|([A-Za-z_][A-Za-z0-9_.]*))!")
EXTERNAL_REF_RE = re.compile(r"\[[^\]]+\]")
HYPERLINK_FN_RE = re.compile(r'HYPERLINK\(\s*"([^"]*)"', re.IGNORECASE)
NA_PATTERN_RE = re.compile(r"\bNA\(\)", re.IGNORECASE)
MAX_SIG_LEN = 400
MAX_TEXT_LEN = 240


def _json_safe(value):
    if isinstance(value, (_dt.datetime, _dt.date, _dt.time)):
        return value.isoformat()
    if isinstance(value, bytes):
        return value.decode("utf-8", "replace")
    return value


def _formula_text(raw) -> str | None:
    """Return formula string for a cell value, handling ArrayFormula objects."""
    if isinstance(raw, str) and raw.startswith("="):
        return raw
    text = getattr(raw, "text", None)  # openpyxl ArrayFormula / DataTableFormula
    if isinstance(text, str) and text.startswith("="):
        return text
    return None


def _color(obj) -> str | None:
    rgb = getattr(obj, "rgb", None)
    return rgb if isinstance(rgb, str) else None


# ---------------------------------------------------------------------------
# Collectors
# ---------------------------------------------------------------------------

def collect_meta(wb_f, path, sheet_filter):
    sheets = []
    for ws in wb_f.worksheets:
        if sheet_filter and ws.title not in sheet_filter:
            continue
        tab = getattr(ws.sheet_properties, "tabColor", None)
        sheets.append({
            "name": ws.title,
            "dims": ws.dimensions,
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "state": ws.sheet_state,           # visible | hidden | veryHidden
            "tab_color": _color(tab),
        })
    return {
        "file": str(path),
        "generator": "extract_workbook.py",
        "sheet_count": len(sheets),
        "sheets": sheets,
    }


def collect_cells_and_friends(wb_f, wb_v, sheet_filter, max_cells):
    """Single pass over every populated cell. Returns (cells, r1c1_rows,
    r1c1_cols, errors, text_inventory, dependencies, truncated_sheets)."""
    cells: dict[str, list] = {}
    r1c1_rows: dict[str, list] = {}
    r1c1_cols: dict[str, list] = {}
    errors: list[dict] = []
    text_inv: dict[str, list] = {}
    edges: dict[tuple, int] = defaultdict(int)
    budget = max_cells
    truncated: list[str] = []

    for ws_f in wb_f.worksheets:
        name = ws_f.title
        if sheet_filter and name not in sheet_filter:
            continue
        ws_v = wb_v[name]
        merged_map = {}
        for rng in ws_f.merged_cells.ranges:
            for coord in expand_range(str(rng)):
                merged_map[coord] = str(rng)

        sheet_cells: list[dict] = []
        sheet_text: list[dict] = []
        row_groups: dict[int, dict] = defaultdict(lambda: defaultdict(list))
        col_groups: dict[int, dict] = defaultdict(lambda: defaultdict(list))
        sheet_truncated = False

        for row_f, row_v in zip(ws_f.iter_rows(), ws_v.iter_rows()):
            for cell_f, cell_v in zip(row_f, row_v):
                raw_f = cell_f.value
                if raw_f is None and cell_v.value is None:
                    continue
                addr = cell_f.coordinate
                r, c = cell_f.row, cell_f.column
                formula = _formula_text(raw_f)
                value = _json_safe(cell_v.value)

                # --- R1C1 pattern groups (always computed, cheap) ---
                if formula:
                    sig = a1_to_r1c1(formula, r, c)[:MAX_SIG_LEN]
                    row_groups[r][sig].append(addr)
                    col_groups[c][sig].append(addr)
                    for m in SHEET_REF_RE.finditer(formula):
                        target = m.group(1) or m.group(2)
                        if EXTERNAL_REF_RE.search(m.group(0)):
                            target = f"[external] {target}"
                        if target and target != name:
                            edges[(name, target)] += 1

                # --- error cells ---
                if isinstance(value, str) and value in ERROR_LITERALS:
                    errors.append({
                        "sheet": name, "addr": addr, "error": value,
                        "formula": formula,
                        "na_pattern": bool(formula and NA_PATTERN_RE.search(formula)),
                    })

                # --- text inventory (labels, headers, narrative) ---
                if (isinstance(value, str) and value not in ERROR_LITERALS
                        and not formula and value.strip()):
                    sheet_text.append({"addr": addr, "text": value[:MAX_TEXT_LEN]})

                # --- full cell record (budgeted) ---
                if budget <= 0:
                    sheet_truncated = True
                    continue
                budget -= 1
                rec = {"addr": addr, "r": r, "c": c}
                if value is not None:
                    rec["v"] = value
                if formula:
                    rec["f"] = formula
                    rec["r1c1"] = a1_to_r1c1(formula, r, c)[:MAX_SIG_LEN]
                fmt = cell_f.number_format
                if fmt and fmt != "General":
                    rec["fmt"] = fmt
                try:
                    style = cell_f.style
                    if isinstance(style, str) and style != "Normal":
                        rec["style"] = style
                except Exception:
                    pass
                font = cell_f.font
                font_rec = {}
                if font.bold:
                    font_rec["b"] = True
                if font.italic:
                    font_rec["i"] = True
                fc = _color(font.color)
                if fc and fc not in ("FF000000",):
                    font_rec["color"] = fc
                if font_rec:
                    rec["font"] = font_rec
                fill = cell_f.fill
                if getattr(fill, "patternType", None) == "solid":
                    fg = _color(fill.fgColor)
                    if fg and fg != "00000000":
                        rec["fill"] = fg
                align = cell_f.alignment.horizontal
                if align:
                    rec["align"] = align
                border = cell_f.border
                edges_str = "".join(s for s, e in zip(
                    "TBLR", (border.top, border.bottom, border.left, border.right))
                    if e is not None and e.style)
                if edges_str:
                    rec["borders"] = edges_str
                if (r, c) in merged_map:
                    rec["merged"] = merged_map[(r, c)]
                if cell_f.comment is not None:
                    rec["comment"] = str(cell_f.comment.text)[:MAX_TEXT_LEN]
                sheet_cells.append(rec)

        cells[name] = sheet_cells
        text_inv[name] = sheet_text
        r1c1_rows[name] = [
            {"row": r, "n_patterns": len(groups),
             "patterns": {sig: addrs for sig, addrs in groups.items()}}
            for r, groups in sorted(row_groups.items()) if groups
        ]
        r1c1_cols[name] = [
            {"col": col_letter(c), "n_patterns": len(groups),
             "patterns": {sig: addrs for sig, addrs in groups.items()}}
            for c, groups in sorted(col_groups.items()) if groups
        ]
        if sheet_truncated:
            truncated.append(name)

    dependencies = {
        "sheet_edges": [
            {"from": a, "to": b, "count": n}
            for (a, b), n in sorted(edges.items(), key=lambda kv: -kv[1])
        ]
    }
    return cells, r1c1_rows, r1c1_cols, errors, text_inv, dependencies, truncated


def collect_named_ranges(wb_f):
    out = []

    def _add(name, dn, scope):
        value = getattr(dn, "value", None) or getattr(dn, "attr_text", "") or ""
        out.append({
            "name": name, "refers_to": value, "scope": scope,
            "broken": "#REF!" in value,
            "hidden": bool(getattr(dn, "hidden", False)),
        })

    dns = wb_f.defined_names
    try:
        items = list(dns.items())          # openpyxl >= 3.1 (dict-like)
    except AttributeError:                  # openpyxl 3.0.x
        items = [(dn.name, dn) for dn in dns.definedName]
    for name, dn in items:
        local = getattr(dn, "localSheetId", None)
        scope = wb_f.sheetnames[local] if isinstance(local, int) else "workbook"
        _add(name, dn, scope)
    # Sheet-scoped names (3.1 stores them on the worksheet)
    for ws in wb_f.worksheets:
        ws_dns = getattr(ws, "defined_names", None)
        if ws_dns:
            try:
                for name, dn in ws_dns.items():
                    _add(name, dn, ws.title)
            except AttributeError:
                pass
    return out


def collect_validations(wb_f, sheet_filter, sheet_names, defined):
    out = []
    for ws in wb_f.worksheets:
        if sheet_filter and ws.title not in sheet_filter:
            continue
        for dv in ws.data_validations.dataValidation:
            f1 = dv.formula1 or ""
            broken = "#REF!" in f1
            m = SHEET_REF_RE.search(f1)
            if m:
                tgt = m.group(1) or m.group(2)
                if tgt not in sheet_names:
                    broken = True
            elif f1.startswith("=") and re.fullmatch(r"=[A-Za-z_][A-Za-z0-9_.]*", f1):
                if f1[1:] not in defined:
                    broken = True
            out.append({
                "sheet": ws.title, "sqref": str(dv.sqref), "type": dv.type,
                "operator": dv.operator, "formula1": dv.formula1,
                "formula2": dv.formula2, "broken_source": broken,
            })
    return out


def collect_conditional_formatting(wb_f, sheet_filter):
    out = []
    for ws in wb_f.worksheets:
        if sheet_filter and ws.title not in sheet_filter:
            continue
        try:
            cf_list = list(ws.conditional_formatting)
        except Exception:
            continue
        for cf in cf_list:
            for rule in cf.rules:
                out.append({
                    "sheet": ws.title, "range": str(cf.sqref),
                    "type": rule.type, "operator": getattr(rule, "operator", None),
                    "formulas": [str(f) for f in (rule.formula or [])],
                    "priority": rule.priority,
                })
    return out


def collect_hyperlinks(wb_f, sheet_filter, sheet_names, defined):
    out = []

    def _internal_broken(location: str) -> bool:
        loc = location.lstrip("#")
        if "!" in loc:
            sheet = loc.split("!")[0].strip("'")
            return sheet not in sheet_names
        return loc not in defined and loc.strip("'") not in sheet_names

    for ws in wb_f.worksheets:
        if sheet_filter and ws.title not in sheet_filter:
            continue
        for hl in getattr(ws, "_hyperlinks", []):
            location = hl.location or ""
            target = hl.target or ""
            internal = bool(location) and not target
            out.append({
                "sheet": ws.title, "addr": hl.ref,
                "target": target or location, "display": hl.display,
                "internal": internal, "source": "ui",
                "broken": _internal_broken(location) if internal else False,
            })
        # HYPERLINK() formulas — invisible to ws._hyperlinks
        for row in ws.iter_rows():
            for cell in row:
                formula = _formula_text(cell.value)
                if not formula or "HYPERLINK" not in formula.upper():
                    continue
                m = HYPERLINK_FN_RE.search(formula)
                tgt = m.group(1) if m else "(dynamic)"
                internal = tgt.startswith("#")
                out.append({
                    "sheet": ws.title, "addr": cell.coordinate,
                    "target": tgt, "display": None,
                    "internal": internal, "source": "formula",
                    "broken": _internal_broken(tgt) if internal else False,
                })
    return out


def collect_styles(wb_f):
    names = []
    try:
        for st in wb_f.named_styles:
            names.append(st if isinstance(st, str) else getattr(st, "name", str(st)))
    except Exception:
        pass
    return [{"name": n, "custom": n not in BUILTIN_STYLES} for n in names]


def collect_charts(wb_f, sheet_filter):
    out = []
    for ws in wb_f.worksheets:
        if sheet_filter and ws.title not in sheet_filter:
            continue
        for chart in getattr(ws, "_charts", []):
            title = None
            try:
                rich = chart.title.tx.rich
                title = "".join(r.t or "" for p in rich.p for r in (p.r or []))
            except Exception:
                pass
            series_ranges = []
            try:
                for ser in chart.series:
                    for ref_holder in (getattr(ser, "val", None), getattr(ser, "cat", None)):
                        for attr in ("numRef", "strRef"):
                            ref = getattr(ref_holder, attr, None)
                            if ref is not None and getattr(ref, "f", None):
                                series_ranges.append(ref.f)
            except Exception:
                pass
            out.append({"sheet": ws.title, "title": title, "series_ranges": series_ranges})
    return out


def flag_chart_fed_errors(errors, charts):
    """Mark error cells whose address sits inside a chart series range —
    candidates for the intentional #N/A chart-gap pattern."""
    ranges_by_sheet: dict[str, list] = defaultdict(list)
    for ch in charts:
        for f in ch["series_ranges"]:
            for part in f.split(","):
                m = re.match(r"(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_. ]*))!(.+)", part.strip())
                if not m:
                    continue
                sheet = m.group(1) or m.group(2)
                try:
                    ranges_by_sheet[sheet].append(expand_range(m.group(3).replace("$", "")))
                except ValueError:
                    pass
    for err in errors:
        coords = None
        try:
            coords = next(iter(expand_range(err["addr"])))
        except ValueError:
            pass
        err["in_chart_range"] = any(
            coords in rng for rng in ranges_by_sheet.get(err["sheet"], [])
        ) if coords else False
    return errors


# ---------------------------------------------------------------------------
# Digest (human-readable companion)
# ---------------------------------------------------------------------------

def write_digest(model: dict, path: str) -> None:
    out = [f"# Workbook extract digest: {Path(model['meta']['file']).name}", ""]
    out.append("## Sheets")
    out.append("| # | Sheet | Dims | State |")
    out.append("|---|-------|------|-------|")
    for i, s in enumerate(model["meta"]["sheets"], 1):
        out.append(f"| {i} | {s['name']} | {s['dims']} | {s['state']} |")
    nr = model["named_ranges"]
    out.append(f"\n## Named ranges: {len(nr)} ({sum(1 for n in nr if n['broken'])} broken)")
    errs = model["errors"]
    out.append(f"## Error cells: {len(errs)}")
    for e in errs[:30]:
        flags = " (chart-fed)" if e.get("in_chart_range") else ""
        out.append(f"- {e['sheet']}!{e['addr']}: {e['error']}{flags}")
    breaks = sum(
        1 for sheet_rows in model["r1c1_rows"].values()
        for rec in sheet_rows if rec["n_patterns"] > 1)
    out.append(f"\n## Rows with >1 R1C1 pattern (break candidates): {breaks}")
    out.append(f"## Validations: {len(model['validations'])}"
               f" | CF rules: {len(model['conditional_formatting'])}"
               f" | Hyperlinks: {len(model['hyperlinks'])}"
               f" | Charts: {len(model['charts'])}")
    customs = [s["name"] for s in model["styles"] if s["custom"]]
    out.append(f"## Custom cell styles ({len(customs)}): {', '.join(customs) or '(none)'}")
    edges = model["dependencies"]["sheet_edges"]
    out.append(f"\n## Cross-sheet dependency edges ({len(edges)})")
    for e in edges[:40]:
        out.append(f"- {e['from']} -> {e['to']} ({e['count']} refs)")
    Path(path).write_text("\n".join(out) + "\n", encoding="utf-8")


# ---------------------------------------------------------------------------

def extract(path: str, sheets: list[str] | None = None,
            max_cells: int = 200000) -> dict:
    wb_f = load_workbook(path, data_only=False)
    wb_v = load_workbook(path, data_only=True)
    sheet_filter = set(sheets) if sheets else None
    sheet_names = set(wb_f.sheetnames)

    meta = collect_meta(wb_f, path, sheet_filter)
    (cells, r1c1_rows, r1c1_cols, errors, text_inv,
     dependencies, truncated) = collect_cells_and_friends(
        wb_f, wb_v, sheet_filter, max_cells)
    named = collect_named_ranges(wb_f)
    defined = {n["name"] for n in named}
    charts = collect_charts(wb_f, sheet_filter)

    model = {
        "meta": meta,
        "cells": cells,
        "r1c1_rows": r1c1_rows,
        "r1c1_cols": r1c1_cols,
        "errors": flag_chart_fed_errors(errors, charts),
        "named_ranges": named,
        "validations": collect_validations(wb_f, sheet_filter, sheet_names, defined),
        "conditional_formatting": collect_conditional_formatting(wb_f, sheet_filter),
        "hyperlinks": collect_hyperlinks(wb_f, sheet_filter, sheet_names, defined),
        "text_inventory": text_inv,
        "dependencies": dependencies,
        "styles": collect_styles(wb_f),
        "charts": charts,
    }
    if truncated:
        model["meta"]["truncated_sheets"] = truncated
    return model


def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("path", help="Path to .xlsx/.xlsm")
    ap.add_argument("--out", default="extract.json", help="Output JSON path")
    ap.add_argument("--sheets", help="Comma-separated sheet names (default: all)")
    ap.add_argument("--digest", help="Also write a human-readable markdown digest")
    ap.add_argument("--max-cells", type=int, default=200000,
                    help="Global cap on per-cell records (pattern/error scans always run on everything)")
    args = ap.parse_args(argv)

    if not Path(args.path).exists():
        sys.exit(f"File not found: {args.path}")
    sheets = [s.strip() for s in args.sheets.split(",")] if args.sheets else None
    model = extract(args.path, sheets=sheets, max_cells=args.max_cells)

    with open(args.out, "w", encoding="utf-8") as fh:
        json.dump(model, fh, indent=1, ensure_ascii=False, default=str)
    n_cells = sum(len(v) for v in model["cells"].values())
    print(f"Extracted {model['meta']['sheet_count']} sheets, {n_cells} cell records, "
          f"{len(model['errors'])} error cells, {len(model['named_ranges'])} names "
          f"-> {args.out}")
    if model["meta"].get("truncated_sheets"):
        print(f"NOTE: cell records truncated on: {', '.join(model['meta']['truncated_sheets'])} "
              f"(raise --max-cells to capture all)")
    if args.digest:
        write_digest(model, args.digest)
        print(f"Digest -> {args.digest}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
