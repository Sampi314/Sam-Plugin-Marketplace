#!/usr/bin/env python3
"""
Excel AI Auditor Scanner
========================
Detects common errors introduced by AI tools (Claude, GPT, Copilot, etc.)
when building Excel financial models via openpyxl, pywin32, or similar.

Usage:
    python ai_audit_scanner.py <filepath> [--sheet SHEET_NAME] [--output OUTPUT_PATH]

Output: Markdown report written to stdout or to the specified output path.
"""

import argparse
import sys
import re
from pathlib import Path
from datetime import datetime, date
from collections import defaultdict, Counter

# Windows consoles default to cp1252, which cannot print the severity emoji.
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, ValueError):
        pass

try:
    import openpyxl
    from openpyxl.utils import get_column_letter, column_index_from_string
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl --break-system-packages")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

CALCULATION_KEYWORDS = {
    "total", "subtotal", "sum", "net", "gross", "balance", "variance",
    "growth", "change", "ratio", "margin", "ebitda", "ebit", "npat",
    "nopat", "fcf", "fcfe", "fcff", "dscr", "llcr", "plcr", "icr",
    "irr", "npv", "wacc", "roic", "roe", "roa", "eps", "pe",
    "profit", "loss", "income", "expense", "depreciation", "amortisation",
    "amortization", "interest", "tax", "revenue", "cost", "cogs",
    "opex", "capex", "closing", "opening", "average", "mean",
    "weighted", "cumulative", "running", "check", "difference",
    "delta", "movement", "cashflow", "cash flow",
}

INPUT_KEYWORDS = {
    "input", "assumption", "constant", "rate", "parameter", "scenario",
    "switch", "flag", "toggle", "driver", "given", "provided",
}

DATE_PATTERNS = [
    r"\d{4}-\d{2}-\d{2}",        # 2024-01-01
    r"\d{2}/\d{2}/\d{4}",        # 01/01/2024
    r"\d{2}-\d{2}-\d{4}",        # 01-01-2024
    r"(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{4}",  # Jan 2024
    r"\d{1,2}\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{4}",  # 1 Jan 2024
]

PERCENTAGE_KEYWORDS = {"rate", "margin", "%", "percentage", "growth", "escalation", "yield", "return"}
CURRENCY_KEYWORDS = {"$", "revenue", "cost", "price", "balance", "cash", "fee", "payment", "salary", "wage"}
RATIO_KEYWORDS = {"ratio", "multiple", "dscr", "llcr", "icr", "plcr", "x"}


# ---------------------------------------------------------------------------
# Finding dataclass
# ---------------------------------------------------------------------------

class Finding:
    def __init__(self, sheet, cell_ref, location_desc, category, description, severity):
        self.sheet = sheet
        self.cell_ref = cell_ref
        self.location_desc = location_desc
        self.category = category
        self.description = description
        self.severity = severity  # "CRITICAL", "HIGH", "MEDIUM", "LOW"

    @property
    def severity_prefix(self):
        prefixes = {
            "CRITICAL": "🔴 CRITICAL:",
            "HIGH": "🔴 HIGH:",
            "MEDIUM": "⚠️ MEDIUM:",
            "LOW": "🟡 LOW:",
        }
        return prefixes.get(self.severity, "")


# ---------------------------------------------------------------------------
# Utility functions
# ---------------------------------------------------------------------------

def get_row_label(ws, row, label_cols=None):
    """Get the text label for a row from the leftmost columns."""
    if label_cols is None:
        label_cols = [1, 2, 3]
    for col in label_cols:
        val = ws.cell(row=row, column=col).value
        if val and isinstance(val, str) and val.strip():
            return val.strip()
    return ""


def get_col_header(ws, col, header_rows=None):
    """Get the column header text."""
    if header_rows is None:
        header_rows = [1, 2, 3]
    for row in header_rows:
        val = ws.cell(row=row, column=col).value
        if val and isinstance(val, str) and val.strip():
            return val.strip()
    return ""


def label_implies_calculation(label):
    """Check if a row label suggests the cell should contain a formula."""
    if not label:
        return False
    label_lower = label.lower().strip()
    for kw in CALCULATION_KEYWORDS:
        if kw in label_lower:
            return True
    return False


def label_implies_input(label):
    """Check if a row label suggests the cell is an input/assumption."""
    if not label:
        return False
    label_lower = label.lower().strip()
    for kw in INPUT_KEYWORDS:
        if kw in label_lower:
            return True
    return False


def is_numeric_string(value):
    """Check if a string value could be parsed as a number."""
    if not isinstance(value, str):
        return False
    cleaned = value.strip().replace(",", "").replace("$", "").replace("£", "").replace("€", "")
    cleaned = cleaned.rstrip("%")
    try:
        float(cleaned)
        return True
    except (ValueError, TypeError):
        return False


def is_date_string(value):
    """Check if a string looks like a date."""
    if not isinstance(value, str):
        return False
    for pattern in DATE_PATTERNS:
        if re.match(pattern, value.strip(), re.IGNORECASE):
            return True
    return False


def extract_sheet_refs_from_formula(formula):
    """Extract sheet names referenced in a formula."""
    if not isinstance(formula, str):
        return []
    # Match 'Sheet Name'!  or  SheetName!
    quoted = re.findall(r"'([^']+)'!", formula)
    unquoted = re.findall(r"(?<!')(\w+)!", formula)
    return quoted + unquoted


def extract_cell_refs_from_formula(formula):
    """Extract cell references from a formula (simplified)."""
    if not isinstance(formula, str):
        return []
    # Match A1-style references, optionally with sheet prefix
    refs = re.findall(r"(?:'[^']*'!|\w+!)?\$?[A-Z]{1,3}\$?\d+", formula, re.IGNORECASE)
    return refs


def find_data_start_col(ws, header_rows=None):
    """Find the first column that contains data (past the label columns)."""
    if header_rows is None:
        header_rows = [1, 2, 3]
    for col in range(1, ws.max_column + 1):
        for row in header_rows:
            val = ws.cell(row=row, column=col).value
            if val and isinstance(val, str):
                val_str = val.strip()
                # Look for year-like, quarter-like, or month-like headers
                if re.match(r"(FY|CY|Q[1-4]|20\d{2}|19\d{2}|Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)", val_str, re.IGNORECASE):
                    return col
    # Fallback: assume column 3+ is data
    return min(4, ws.max_column)


def find_header_row(ws):
    """Heuristic to find the header row (the row with the most text values)."""
    best_row = 1
    best_count = 0
    for row in range(1, min(10, ws.max_row + 1)):
        count = sum(1 for col in range(1, ws.max_column + 1)
                    if ws.cell(row=row, column=col).value and isinstance(ws.cell(row=row, column=col).value, str))
        if count > best_count:
            best_count = count
            best_row = row
    return best_row


# ---------------------------------------------------------------------------
# Audit checks
# ---------------------------------------------------------------------------

def check_formula_as_text(ws, findings):
    """Phase 2a: Detect formulas stored as text strings."""
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.data_type == 's' and isinstance(cell.value, str) and cell.value.strip().startswith('='):
                cell_ref = f"{get_column_letter(col)}{row}"
                label = get_row_label(ws, row)
                findings.append(Finding(
                    sheet=ws.title,
                    cell_ref=cell_ref,
                    location_desc=f"Row: '{label}'" if label else f"Row {row}",
                    category="Formula-as-Text",
                    description=f"🔴 CRITICAL: Cell contains a formula stored as a text string, not a live formula. "
                                f"The formula '{cell.value[:80]}{'...' if len(cell.value) > 80 else ''}' will NOT calculate. "
                                f"Likely cause: openpyxl wrote the formula as a string value instead of a formula.",
                    severity="CRITICAL",
                ))


def check_static_snapshots(ws, findings):
    """Phase 2b: Detect hard-coded values where formulas are expected."""
    header_row = find_header_row(ws)
    data_start_col = find_data_start_col(ws)

    for row in range(header_row + 1, ws.max_row + 1):
        label = get_row_label(ws, row)
        if not label_implies_calculation(label):
            continue
        if label_implies_input(label):
            continue

        # Check cells in the data area
        has_formula = False
        constant_cells = []
        for col in range(data_start_col, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is None:
                continue
            if isinstance(cell.value, str) and cell.value.startswith('='):
                has_formula = True
            elif isinstance(cell.value, (int, float)):
                constant_cells.append(col)

        # If the row has SOME formulas but also constants, flag the constants
        if has_formula and constant_cells:
            for col in constant_cells:
                cell_ref = f"{get_column_letter(col)}{row}"
                findings.append(Finding(
                    sheet=ws.title,
                    cell_ref=cell_ref,
                    location_desc=f"Row: '{label}'",
                    category="Static Snapshot",
                    description=f"⚠️ MEDIUM: Row '{label}' has formulas in some columns but a hard-coded "
                                f"constant ({ws.cell(row=row, column=col).value}) in this cell. "
                                f"Likely a static snapshot from the AI's Python computation.",
                    severity="MEDIUM",
                ))

        # If the row has NO formulas and only constants, and label implies calculation
        if not has_formula and constant_cells and len(constant_cells) >= 2:
            cell_refs = [f"{get_column_letter(c)}{row}" for c in constant_cells]
            ref_str = f"{cell_refs[0]}:{cell_refs[-1]}" if len(cell_refs) > 1 else cell_refs[0]
            findings.append(Finding(
                sheet=ws.title,
                cell_ref=ref_str,
                location_desc=f"Row: '{label}'",
                category="Static Snapshot",
                description=f"🔴 CRITICAL: Row '{label}' implies a calculation but contains only hard-coded "
                            f"constants across {len(constant_cells)} cells. No formula exists. "
                            f"AI likely computed the values in Python and wrote them as numbers.",
                severity="CRITICAL",
            ))


def check_uniform_values(ws, findings):
    """Phase 2c: Detect rows where all values are identical (AI pasted a variable)."""
    header_row = find_header_row(ws)
    data_start_col = find_data_start_col(ws)

    for row in range(header_row + 1, ws.max_row + 1):
        label = get_row_label(ws, row)
        if label_implies_input(label):
            continue

        values = []
        for col in range(data_start_col, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, (int, float)) and cell.value != 0:
                values.append(cell.value)

        if len(values) >= 3 and len(set(values)) == 1 and label_implies_calculation(label):
            first_col = get_column_letter(data_start_col)
            last_col = get_column_letter(ws.max_column)
            findings.append(Finding(
                sheet=ws.title,
                cell_ref=f"{first_col}{row}:{last_col}{row}",
                location_desc=f"Row: '{label}'",
                category="Uniform Value Fill",
                description=f"⚠️ MEDIUM: Row '{label}' contains the identical value ({values[0]}) across "
                            f"{len(values)} period columns. This suggests the AI wrote a single computed "
                            f"value into every cell rather than building period-specific formulas.",
                severity="MEDIUM",
            ))


def check_text_as_number(ws, findings):
    """Phase 4a: Detect numeric values stored as text strings."""
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.data_type == 's' and is_numeric_string(cell.value):
                # Skip if it looks like a formula-as-text (caught elsewhere)
                if cell.value.strip().startswith('='):
                    continue
                cell_ref = f"{get_column_letter(col)}{row}"
                label = get_row_label(ws, row)
                findings.append(Finding(
                    sheet=ws.title,
                    cell_ref=cell_ref,
                    location_desc=f"Row: '{label}'" if label else f"Row {row}",
                    category="Text-as-Number",
                    description=f"🔴 CRITICAL: Cell contains the numeric value '{cell.value}' stored as a TEXT "
                                f"string. This value will be INVISIBLE to SUM, AVERAGE, and other numeric "
                                f"functions. Likely cause: openpyxl wrote a string instead of int/float.",
                    severity="CRITICAL",
                ))


def check_date_as_text(ws, findings):
    """Phase 4b: Detect dates stored as text strings."""
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.data_type == 's' and is_date_string(cell.value):
                cell_ref = f"{get_column_letter(col)}{row}"
                label = get_row_label(ws, row)
                findings.append(Finding(
                    sheet=ws.title,
                    cell_ref=cell_ref,
                    location_desc=f"Row: '{label}'" if label else f"Row {row}",
                    category="Date-as-Text",
                    description=f"⚠️ MEDIUM: Cell contains a date-like value '{cell.value}' stored as text. "
                                f"Date arithmetic and sorting will not work correctly. "
                                f"Should be a datetime object or Excel serial number.",
                    severity="MEDIUM",
                ))


def check_boolean_as_text(ws, findings):
    """Phase 4c: Detect TRUE/FALSE stored as text."""
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.data_type == 's' and isinstance(cell.value, str) and cell.value.strip().upper() in ("TRUE", "FALSE"):
                cell_ref = f"{get_column_letter(col)}{row}"
                label = get_row_label(ws, row)
                findings.append(Finding(
                    sheet=ws.title,
                    cell_ref=cell_ref,
                    location_desc=f"Row: '{label}'" if label else f"Row {row}",
                    category="Boolean-as-Text",
                    description=f"🟡 LOW: Cell contains '{cell.value}' stored as text instead of a boolean value. "
                                f"Logical formulas (IF, AND, OR) may not evaluate correctly.",
                    severity="LOW",
                ))


def check_empty_placeholders(ws, findings):
    """Phase 1b: Detect rows with labels but no data."""
    data_start_col = find_data_start_col(ws)
    header_row = find_header_row(ws)

    for row in range(header_row + 1, ws.max_row + 1):
        label = get_row_label(ws, row)
        if not label:
            continue

        # Skip obvious spacer/separator labels
        label_lower = label.lower().strip()
        if label_lower in ("", " ") or all(c in "-=_ " for c in label_lower):
            continue

        has_any_data = False
        for col in range(data_start_col, ws.max_column + 1):
            if ws.cell(row=row, column=col).value is not None:
                has_any_data = True
                break

        if not has_any_data and label_implies_calculation(label):
            findings.append(Finding(
                sheet=ws.title,
                cell_ref=f"{get_column_letter(data_start_col)}{row}:{get_column_letter(ws.max_column)}{row}",
                location_desc=f"Row: '{label}'",
                category="Empty Placeholder",
                description=f"🔴 HIGH: Row '{label}' has a label implying a calculation but all data "
                            f"cells are empty. The AI created the structure but forgot to write "
                            f"the formulas.",
                severity="HIGH",
            ))


def check_broken_sheet_refs(ws, wb_sheetnames, findings):
    """Phase 1a: Detect formula references to non-existent sheets."""
    sheetnames_lower = {s.lower(): s for s in wb_sheetnames}
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.data_type == 'f' or (isinstance(cell.value, str) and cell.value.startswith('=')):
                formula = cell.value if isinstance(cell.value, str) else ""
                sheet_refs = extract_sheet_refs_from_formula(formula)
                for ref_sheet in sheet_refs:
                    if ref_sheet.lower() not in sheetnames_lower:
                        cell_ref = f"{get_column_letter(col)}{row}"
                        label = get_row_label(ws, row)
                        findings.append(Finding(
                            sheet=ws.title,
                            cell_ref=cell_ref,
                            location_desc=f"Row: '{label}'" if label else f"Row {row}",
                            category="Broken Sheet Link",
                            description=f"🔴 HIGH: Formula references sheet '{ref_sheet}' which does not "
                                        f"exist in the workbook. Available sheets: {', '.join(wb_sheetnames)}.",
                            severity="HIGH",
                        ))


def check_sum_boundaries(ws, findings):
    """Phase 3a: Check SUM formulas for off-by-one boundary errors."""
    header_row = find_header_row(ws)

    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            val = cell.value
            if not isinstance(val, str) or not val.startswith('='):
                continue

            # Find SUM ranges
            sum_matches = re.findall(r'SUM\(([A-Z]+\d+):([A-Z]+\d+)\)', val, re.IGNORECASE)
            for start_ref, end_ref in sum_matches:
                try:
                    start_col_letter = re.match(r'([A-Z]+)', start_ref, re.IGNORECASE).group(1)
                    start_row_num = int(re.search(r'(\d+)', start_ref).group(1))
                    end_row_num = int(re.search(r'(\d+)', end_ref).group(1))
                except (AttributeError, ValueError):
                    continue

                cell_ref = f"{get_column_letter(col)}{row}"
                label = get_row_label(ws, row)

                # Check if SUM includes the header row
                if start_row_num <= header_row:
                    findings.append(Finding(
                        sheet=ws.title,
                        cell_ref=cell_ref,
                        location_desc=f"Row: '{label}'" if label else f"Row {row}",
                        category="SUM Boundary Error",
                        description=f"🔴 HIGH: SUM formula includes header row {header_row}. "
                                    f"Range {start_ref}:{end_ref} starts at row {start_row_num} "
                                    f"but the header appears to be at row {header_row}.",
                        severity="HIGH",
                    ))

                # Check if SUM includes its own row (circular)
                if start_row_num <= row <= end_row_num:
                    findings.append(Finding(
                        sheet=ws.title,
                        cell_ref=cell_ref,
                        location_desc=f"Row: '{label}'" if label else f"Row {row}",
                        category="SUM Boundary Error",
                        description=f"🔴 HIGH: SUM formula {start_ref}:{end_ref} includes its own cell "
                                    f"at row {row}, creating a circular reference.",
                        severity="HIGH",
                    ))


def check_missing_error_handling(ws, findings):
    """Phase 6a: Check for unprotected lookups and divisions."""
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            val = cell.value
            if not isinstance(val, str) or not val.startswith('='):
                continue

            cell_ref = f"{get_column_letter(col)}{row}"
            label = get_row_label(ws, row)
            val_upper = val.upper()

            # Check for lookup functions without IFERROR
            lookup_funcs = ["VLOOKUP", "HLOOKUP", "XLOOKUP", "INDEX", "MATCH"]
            for func in lookup_funcs:
                if func in val_upper and "IFERROR" not in val_upper and "IFNA" not in val_upper:
                    findings.append(Finding(
                        sheet=ws.title,
                        cell_ref=cell_ref,
                        location_desc=f"Row: '{label}'" if label else f"Row {row}",
                        category="Missing Error Handling",
                        description=f"🟡 LOW: {func} formula lacks IFERROR/IFNA wrapper. "
                                    f"Will show #N/A or #REF! if lookup value is not found.",
                        severity="LOW",
                    ))
                    break  # One finding per cell

            # Check for division without protection
            if "/" in val and "IFERROR" not in val_upper and "IF(" not in val_upper:
                # Quick heuristic: is there a raw division?
                # Exclude date-like patterns (1/1/2024)
                div_parts = re.findall(r'[A-Z]+\d+\s*/\s*[A-Z]+\d+', val, re.IGNORECASE)
                if div_parts:
                    # Check if divisor could be zero
                    findings.append(Finding(
                        sheet=ws.title,
                        cell_ref=cell_ref,
                        location_desc=f"Row: '{label}'" if label else f"Row {row}",
                        category="Missing Error Handling",
                        description=f"🟡 LOW: Division formula without divide-by-zero protection. "
                                    f"Consider wrapping in IFERROR or adding IF(denominator=0,...) guard.",
                        severity="LOW",
                    ))


def check_number_format_gaps(ws, findings):
    """Phase 5a: Check for missing number formats."""
    header_row = find_header_row(ws)

    for row in range(header_row + 1, ws.max_row + 1):
        label = get_row_label(ws, row)
        if not label:
            continue
        label_lower = label.lower()

        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is None:
                continue
            if not isinstance(cell.value, (int, float)):
                continue

            fmt = cell.number_format or "General"
            cell_ref = f"{get_column_letter(col)}{row}"

            # Check percentages
            if any(kw in label_lower for kw in PERCENTAGE_KEYWORDS):
                if "%" not in fmt and fmt == "General":
                    if isinstance(cell.value, (int, float)) and -2 < cell.value < 2:
                        findings.append(Finding(
                            sheet=ws.title,
                            cell_ref=cell_ref,
                            location_desc=f"Row: '{label}'",
                            category="Number Format Gap",
                            description=f"🟡 LOW: Row label '{label}' implies a percentage but the cell "
                                        f"has General format. Value {cell.value} likely represents "
                                        f"{cell.value*100:.1f}% but will display as a raw decimal.",
                            severity="LOW",
                        ))


def check_absolute_relative_refs(ws, findings):
    """Phase 3b: Check for over-use of absolute references in time-series."""
    header_row = find_header_row(ws)
    data_start_col = find_data_start_col(ws)

    for row in range(header_row + 1, ws.max_row + 1):
        formulas = {}
        for col in range(data_start_col, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, str) and cell.value.startswith('='):
                formulas[col] = cell.value

        if len(formulas) < 3:
            continue

        # Check if all formulas are literally identical (strong sign of absolute-everywhere)
        unique_formulas = set(formulas.values())
        if len(unique_formulas) == 1:
            label = get_row_label(ws, row)
            if label and not label_implies_input(label):
                cols = sorted(formulas.keys())
                first = get_column_letter(cols[0])
                last = get_column_letter(cols[-1])
                findings.append(Finding(
                    sheet=ws.title,
                    cell_ref=f"{first}{row}:{last}{row}",
                    location_desc=f"Row: '{label}'",
                    category="Reference Type Error",
                    description=f"⚠️ MEDIUM: All {len(formulas)} formulas in row '{label}' are literally identical "
                                f"('{list(unique_formulas)[0][:60]}'). In a time-series, formulas should adjust "
                                f"across columns. AI likely used all-absolute references.",
                    severity="MEDIUM",
                ))


def check_print_setup(wb, findings):
    """Phase 6d: Check if print areas and page setup are configured."""
    any_print_area = False
    for ws in wb.worksheets:
        if ws.print_area:
            any_print_area = True
            break

    if not any_print_area:
        findings.append(Finding(
            sheet="(Workbook)",
            cell_ref="N/A",
            location_desc="Workbook-level setting",
            category="Missing Print Setup",
            description="🟡 LOW: No print areas are defined in any sheet. AI tools almost never set "
                        "print areas, page orientation, scaling, or margins.",
            severity="LOW",
        ))


def check_named_ranges(wb, findings):
    """Phase 6b: Check named range hygiene."""
    defined_names = list(wb.defined_names.definedName) if hasattr(wb.defined_names, 'definedName') else []

    for dn in defined_names:
        ref = str(dn.attr_text) if hasattr(dn, 'attr_text') else str(dn.value)
        name = dn.name

        if "#REF" in ref:
            findings.append(Finding(
                sheet="(Name Manager)",
                cell_ref=name,
                location_desc=f"Named Range: '{name}'",
                category="Dead Named Range",
                description=f"🔴 HIGH: Named range '{name}' points to #REF! ({ref}). "
                            f"The referenced cells or sheet have been deleted.",
                severity="HIGH",
            ))


# ---------------------------------------------------------------------------
# Main orchestrator
# ---------------------------------------------------------------------------

def run_audit(filepath, target_sheets=None):
    """Run the full AI audit on an Excel file."""
    wb = openpyxl.load_workbook(filepath, data_only=False)
    findings = []

    sheets_to_audit = []
    for ws in wb.worksheets:
        if target_sheets and ws.title not in target_sheets:
            continue
        sheets_to_audit.append(ws)

    wb_sheetnames = [ws.title for ws in wb.worksheets]

    # Run all checks
    for ws in sheets_to_audit:
        if ws.max_row is None or ws.max_row < 1:
            continue
        check_formula_as_text(ws, findings)
        check_static_snapshots(ws, findings)
        check_uniform_values(ws, findings)
        check_text_as_number(ws, findings)
        check_date_as_text(ws, findings)
        check_boolean_as_text(ws, findings)
        check_empty_placeholders(ws, findings)
        check_broken_sheet_refs(ws, wb_sheetnames, findings)
        check_sum_boundaries(ws, findings)
        check_missing_error_handling(ws, findings)
        check_number_format_gaps(ws, findings)
        check_absolute_relative_refs(ws, findings)

    # Workbook-level checks
    check_print_setup(wb, findings)
    check_named_ranges(wb, findings)

    wb.close()
    return findings


# Unified severity vocabulary (audit_standards.md §3): the scanner's internal
# CRITICAL/HIGH/MEDIUM/LOW detection grades map onto Critical/Warning/Info.
SEVERITY_DISPLAY = {
    "CRITICAL": "🔴 Critical",
    "HIGH": "🔴 Critical",
    "MEDIUM": "⚠️ Warning",
    "LOW": "🟡 Info",
}

# Legacy severity prefixes embedded in detection descriptions — stripped at
# emission time now that Severity is a dedicated column.
_LEGACY_PREFIX_RE = re.compile(r"^(?:🔴 CRITICAL:|🔴 HIGH:|⚠️ MEDIUM:|🟡 LOW:)\s*")


def generate_report(findings, filepath):
    """Generate the markdown report."""
    # Count statistics
    categories = Counter(f.category for f in findings)
    severities = Counter(f.severity for f in findings)

    critical_cats = {"Formula-as-Text", "Static Snapshot", "Text-as-Number", "Broken Sheet Link"}
    structural_cats = {"Empty Placeholder", "Orphaned Reference", "SUM Boundary Error",
                       "Uniform Value Fill", "Reference Type Error"}
    formatting_cats = {"Number Format Gap", "Formatting Inconsistency", "Missing Print Setup"}

    critical_count = sum(1 for f in findings if f.category in critical_cats)
    structural_count = sum(1 for f in findings if f.category in structural_cats)
    formatting_count = sum(1 for f in findings if f.category in formatting_cats)

    # Determine overall rating
    formula_as_text_count = categories.get("Formula-as-Text", 0)
    static_snapshot_count = categories.get("Static Snapshot", 0)

    if formula_as_text_count > 0 or critical_count > 5:
        overall = "🔴 CRITICAL FAIL"
    elif critical_count > 0 or structural_count > 5:
        overall = "⚠️ FAIL"
    elif structural_count > 0 or formatting_count > 10:
        overall = "🟡 MARGINAL PASS"
    else:
        overall = "✅ PASS"

    lines = []
    lines.append(f"# AI Audit Report: {Path(filepath).name}")
    lines.append(f"*Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}*")
    lines.append("")
    lines.append("```")
    lines.append("AI Audit Summary")
    lines.append("═" * 50)
    lines.append(f"  Total findings:               {len(findings)}")
    lines.append(f"")
    lines.append(f"  Critical AI Errors:            {critical_count}")
    lines.append(f"    Formula-as-text:             {categories.get('Formula-as-Text', 0)}")
    lines.append(f"    Static snapshots:            {categories.get('Static Snapshot', 0)}")
    lines.append(f"    Text-as-number:              {categories.get('Text-as-Number', 0)}")
    lines.append(f"    Broken sheet references:     {categories.get('Broken Sheet Link', 0)}")
    lines.append(f"")
    lines.append(f"  Structural Issues:             {structural_count}")
    lines.append(f"    Empty placeholder rows:      {categories.get('Empty Placeholder', 0)}")
    lines.append(f"    SUM boundary errors:         {categories.get('SUM Boundary Error', 0)}")
    lines.append(f"    Uniform value fills:         {categories.get('Uniform Value Fill', 0)}")
    lines.append(f"    Reference type errors:       {categories.get('Reference Type Error', 0)}")
    lines.append(f"")
    lines.append(f"  Formatting Gaps:               {formatting_count}")
    lines.append(f"    Missing number formats:      {categories.get('Number Format Gap', 0)}")
    lines.append(f"    Missing print setup:         {categories.get('Missing Print Setup', 0)}")
    lines.append(f"")
    lines.append(f"  Other Issues:                  {len(findings) - critical_count - structural_count - formatting_count}")
    lines.append(f"    Missing error handling:       {categories.get('Missing Error Handling', 0)}")
    lines.append(f"    Dead named ranges:           {categories.get('Dead Named Range', 0)}")
    lines.append(f"    Date-as-text:                {categories.get('Date-as-Text', 0)}")
    lines.append(f"    Boolean-as-text:             {categories.get('Boolean-as-Text', 0)}")
    lines.append(f"")
    lines.append(f"  Overall AI Build Quality:      {overall}")
    lines.append("```")
    lines.append("")

    if not findings:
        lines.append("✅ No issues detected. The model appears to contain live formulas "
                      "and correct data types. Recommend running the standard audit suite (Logic, Sentry, "
                      "Stylist) for business logic validation.")
        return "\n".join(lines)

    # Findings table — unified specialist columns (audit_standards.md §1)
    lines.append("## Findings")
    lines.append("")
    lines.append("| Sheet | Cell Reference | Description of Location | Severity | Category | Description of Issue |")
    lines.append("|---|---|---|---|---|---|")

    # Sort: Critical first, then High, Medium, Low
    severity_order = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3}
    sorted_findings = sorted(findings, key=lambda f: (severity_order.get(f.severity, 9), f.sheet, f.cell_ref))

    for f in sorted_findings:
        severity = SEVERITY_DISPLAY.get(f.severity, f.severity)
        description = _LEGACY_PREFIX_RE.sub("", f.description)
        lines.append(f"| {f.sheet} | {f.cell_ref} | {f.location_desc} | {severity} | {f.category} | {description} |")

    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("*Report generated by the Excel AI Auditor skill.*")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Excel AI Auditor — detect AI-generated model errors")
    parser.add_argument("filepath", help="Path to the Excel file (.xlsx)")
    parser.add_argument("--sheet", nargs="*", help="Specific sheet name(s) to audit (default: all)")
    parser.add_argument("--output", help="Output file path for the report (default: stdout)")
    args = parser.parse_args()

    filepath = Path(args.filepath)
    if not filepath.exists():
        print(f"ERROR: File not found: {filepath}")
        sys.exit(1)
    if not filepath.suffix.lower() in ('.xlsx', '.xlsm'):
        print(f"WARNING: File extension '{filepath.suffix}' is not .xlsx/.xlsm. Attempting anyway...")

    print(f"Scanning: {filepath}", file=sys.stderr)
    findings = run_audit(str(filepath), target_sheets=args.sheet)
    print(f"Found {len(findings)} issues.", file=sys.stderr)

    report = generate_report(findings, str(filepath))

    if args.output:
        out_path = Path(args.output)
        out_path.write_text(report, encoding="utf-8")
        print(f"Report written to: {out_path}", file=sys.stderr)
    else:
        print(report)


if __name__ == "__main__":
    main()
